# === CAD Estimator Pro — main.py (Part 1/4) ==================================
# Importy, konfiguracja, stałe, helpery HTTP/AI i nowe funkcje integracyjne
# ============================================================================

import streamlit as st
import pandas as pd
import psycopg2
from psycopg2.extras import RealDictCursor
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import json
import os
from datetime import datetime
import base64
import re
import logging
from contextlib import contextmanager
import time
from functools import lru_cache
from PIL import Image, ImageDraw, ImageFont
import io
from io import BytesIO
import plotly.express as px
from PyPDF2 import PdfReader
from rapidfuzz import fuzz, process
from openpyxl import load_workbook


# === LOGGING ===
logging.basicConfig(level=logging.DEBUG, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("cad-estimator")

# === UI CONFIG ===
st.set_page_config(page_title="CAD Estimator Pro", layout="wide", page_icon="🚀")

# === ENV ===
OLLAMA_URL = os.getenv('OLLAMA_URL', 'http://ollama:11434')
DB_HOST = os.getenv('DB_HOST', 'cad-postgres')
DB_NAME = os.getenv('DB_NAME', 'cad_estimator')
DB_USER = os.getenv('DB_USER', 'cad_user')
DB_PASSWORD = os.getenv('DB_PASSWORD', 'cad_password_2024')
EMBED_MODEL = os.getenv('EMBED_MODEL', 'nomic-embed-text')
EMBED_DIM = int(os.getenv('EMBED_DIM', '768'))

# === DZIAŁY ===
DEPARTMENTS = {
    '131': 'Automotive',
    '132': 'Industrial Machinery',
    '133': 'Transportation',
    '134': 'Heavy Equipment',
    '135': 'Special Purpose Machinery'
}
DEPARTMENT_CONTEXT = {
    '131': """Branża: AUTOMOTIVE (Faurecia, VW, Merit, Sitech, Joyson)
Specyfika: Komponenty samochodowe, wysokie wymagania jakościowe, spawanie precyzyjne, duże serie produkcyjne, normy automotive (IATF 16949).""",
    '132': """Branża: INDUSTRIAL MACHINERY (PMP, ITM, Amazon)
Specyfika: Maszyny przemysłowe, automatyka, systemy pakowania, linie produkcyjne, robotyka przemysłowa, PLC.""",
    '133': """Branża: TRANSPORTATION (Volvo, Scania)
Specyfika: Pojazdy ciężarowe, autobusy, systemy transportowe, wytrzymałość strukturalna, normy transportowe.""",
    '134': """Branża: HEAVY EQUIPMENT (Volvo CE, Mine Master)
Specyfika: Maszyny budowlane, koparki, ładowarki, ekstremalne obciążenia, odporność na warunki terenowe.""",
    '135': """Branża: SPECIAL PURPOSE MACHINERY (Bosch, Chassis Brakes, BWI, Besta)
Specyfika: Maszyny specjalne, niestandardowe rozwiązania, prototypy, unikalne wymagania klienta."""
}

# === SŁOWNIK NORMALIZACJI KOMPONENTÓW (PL/DE/EN -> EN) ===
# Dodaj więcej aliasów (linia ~52):
COMPONENT_ALIASES = {
    # Wsporniki
    'wspornik': 'bracket', 'wsporniki': 'bracket', 'halterung': 'bracket', 
    'halter': 'bracket', 'träger': 'bracket', 'support': 'bracket', 
    'konsole': 'bracket', 'brackets': 'bracket',
    
    # Ramy
    'rama': 'frame', 'ramy': 'frame', 'rahmen': 'frame', 'gestell': 'frame', 
    'chassis': 'frame', 'frames': 'frame',
    
    # Przenośniki
    'przenośnik': 'conveyor', 'przenośniki': 'conveyor', 'förderband': 'conveyor', 
    'förderer': 'conveyor', 'transport': 'conveyor', 'conveying': 'conveyor',
    'conveyor': 'conveyor',
    
    # Płyty
    'płyta': 'plate', 'płyty': 'plate', 'platte': 'plate', 'platten': 'plate',
    'sheet': 'plate', 'panel': 'plate', 'plates': 'plate',
    
    # Profile
    'profil': 'profile', 'profile': 'profile', 'profiles': 'profile',
    
    # Adaptery
    'adapter': 'adapter', 'adapters': 'adapter', 'adaptador': 'adapter',
    
    # Czujniki
    'czujnik': 'sensor', 'czujniki': 'sensor', 'sensor': 'sensor', 
    'sensors': 'sensor', 'sensoring': 'sensor',
    
    # Prowadnice
    'prowadnica': 'guide', 'prowadnice': 'guide', 'führung': 'guide', 
    'rail': 'guide', 'rails': 'guide', 'guide': 'guide', 'guides': 'guide',
    'guidance': 'guide',
    
    # Łożyska
    'łożysko': 'bearing', 'łożyska': 'bearing', 'lager': 'bearing',
    'bearing': 'bearing', 'bearings': 'bearing',
    
    # Śruby
    'śruba': 'screw', 'śruby': 'screw', 'schraube': 'screw', 
    'screwdriver': 'screw', 'screwdrivers': 'screw',
    
    # Cylindry
    'cylinder': 'cylinder', 'cylinders': 'cylinder', 'zylinder': 'cylinder',
    'siłownik': 'cylinder', 'siłowniki': 'cylinder',
    
    # Bases
    'podstawa': 'base', 'podstawy': 'base', 'basis': 'base', 
    'base': 'base', 'bases': 'base', 'fundament': 'base', 'sockel': 'base',
}
def extract_scope_from_excel_a1_first_sheet(file_like) -> str:
    """
    Zwraca opis z komórki A1 pierwszego arkusza (pierwszej zakładki).
    Obsługuje UploadedFile/bytes/file-like. Zwraca "" jeśli brak.
    """
    try:
        # spłaszcz do bytes
        if hasattr(file_like, "read"):
            content = file_like.read()
        elif isinstance(file_like, bytes):
            content = file_like
        else:
            try:
                pos = file_like.tell()
                file_like.seek(0)
                content = file_like.read()
                file_like.seek(pos)
            except Exception:
                return ""
        wb = load_workbook(BytesIO(content), data_only=True)
        ws = wb.worksheets[0]
        val = ws.cell(row=1, column=1).value
        return str(val).strip() if val is not None else ""
    except Exception as e:
        logger.info(f"Nie udało się odczytać A1 z Excela: {e}")
        return ""
# === PROMPTY ===
MASTER_PROMPT = """Jesteś senior konstruktorem CAD z 20-letnim doświadczeniem w:

Projektowaniu ram spawalniczych i konstrukcji stalowych
Automatyce przemysłowej (PLC, robotyka, pozycjonery)
Systemach wizyjnych i kontroli jakości
Narzędziach CAD: CATIA V5, SolidWorks, AutoCAD
Odpowiadaj ZAWSZE w języku polskim.

# ═══════════════════════════════════════════════════════════
# METODYKA ESTYMACJI - KROK PO KROKU
# ═══════════════════════════════════════════════════════════

1. ANALIZA OPISU (przeczytaj DOKŁADNIE):
   - Zidentyfikuj WSZYSTKIE komponenty wymienione w opisie
   - Sprawdź czy są podane ilości (np. "4x wspornik", "8x otwór")
   - Zwróć uwagę na materiały (S355, S235, aluminum)
   - Zidentyfikuj procesy: spawanie, obróbka, montaż
   - Sprawdź normy (ISO, EN, AWS)

2. WYKRYJ BRAKUJĄCE INFORMACJE:
   Sprawdź czy w opisie jest:
   - ✓ Materiał (S235, S355, aluminium?)
   - ✓ Wymiary (długość, szerokość, grubość?)
   - ✓ Ilości komponentów (4x wspornik?)
   - ✓ Normy (ISO 9606, EN 1090, AWS D1.1?)
   - ✓ Procesy (spawanie MAG/MIG/TIG? obróbka CNC?)
   - ✓ Tolerancje (±0.1mm, ±0.5mm?)
   - ✓ Obróbka powierzchni (malowanie, cynkowanie?)
   
   Jeśli CZEGOKOLWIEK brakuje → wygeneruj pytania w "missing_info"

3. DEKOMPOZYCJA (rozłóż na części):
   - Każdy wymieniony komponent = osobna pozycja w "components"
   - Złożenia = suma części składowych
   - NIE pomijaj żadnego elementu z opisu!

4. ESTYMACJA GODZIN (dla KAŻDEGO komponentu osobno):
   
   LAYOUT (3D koncepcja):
   - Prosta płyta: 0.5-1h
   - Profil/wspornik: 1-2h
   - Złożenie proste: 2-4h
   - Złożenie średnie: 4-8h
   - Złożenie złożone: 8-15h
   
   DETAIL (3D szczegóły):
   - Prosta płyta z otworami: 2-4h
   - Profil spawany: 4-8h
   - Element z obróbką: 5-10h
   - Złożenie z kinematyką: 10-20h
   - Złożenie ze spawami: +20-30%
   
   DOC (dokumentacja 2D):
   - Prosty rysunek: 1-2h
   - Rysunek wykonawczy: 2-4h
   - Złożenie: 3-6h
   - Dokumentacja spawania: +1-2h

5. GENERUJ SUGESTIE:
   Po dekompozycji ZAWSZE sprawdź:
   
   A) CZY MOŻNA ZROBIĆ INACZEJ? (alternative)
      - Spawanie vs śruby vs nitowanie
      - Materiały: stal vs aluminium
      - Procesy: CNC vs laser
   
   B) CZY MOŻNA ULEPSZYĆ? (improvement)
      - Dodać wzmocnienia dla sztywności
      - Zmienić geometrię dla oszczędności
      - Uprościć montaż
   
   C) CZY JEST RYZYKO/OSTRZEŻENIE? (warning)
      - Brak normy spawania → niezgodność
      - Zbyt cienka płyta → odkształcenia
      - Brak wzmocnień → niestabilność

6. WALIDACJA:
   - Suma layout ≈ 15-25% total
   - Suma detail ≈ 50-60% total
   - Suma doc ≈ 20-30% total

# ═══════════════════════════════════════════════════════════
# WYMAGANY FORMAT ODPOWIEDZI - ZWRÓĆ TYLKO CZYSTY JSON
# ═══════════════════════════════════════════════════════════

{
  "components": [
    {"name": "Nazwa DOKŁADNA", "layout_h": 12.5, "detail_h": 42.0, "doc_h": 28.0}
  ],
  "sums": {"layout": 12.5, "detail": 42.0, "doc": 28.0, "total": 82.5},
  "assumptions": ["Założenie 1", "Założenie 2"],
  "risks": [
    {"risk": "Opis ryzyka", "impact": "wysoki/średni/niski", "mitigation": "Jak zminimalizować"}
  ],
  "adjustments": [
    {
      "parent": "Nazwa komponentu",
      "adds": [
        {"name": "sub-komponent", "qty": 2, "layout_add": 0.5, "detail_add": 3.0, "doc_add": 1.0, "reason": "dlaczego"}
      ]
    }
  ],
  "suggestions": [
    {
      "type": "alternative",
      "title": "Krótki tytuł sugestii (max 50 znaków)",
      "description": "Szczegółowy opis (2-3 zdania). Co można zrobić inaczej i jakie będą efekty.",
      "priority": "high",
      "impact": {
        "hours_delta": -5.0,
        "cost_delta": -750,
        "quality_info": "Łatwiejszy demontaż"
      },
      "components_to_add": ["Opcjonalnie: lista komponentów do dodania"],
      "components_to_remove": ["Opcjonalnie: lista do usunięcia"]
    }
  ],
  "missing_info": [
    {
      "question": "Jaki materiał płyty bazowej?",
      "type": "choice",
      "field_name": "material_plate",
      "options": ["S235JR (stal konstrukcyjna)", "S355J2 (wyższa wytrzymałość)", "Aluminium (lżejsze)"],
      "default": "S235JR (stal konstrukcyjna)",
      "priority": "high",
      "why": "Materiał wpływa na czas obróbki (+20% dla S355) i koszty"
    },
    {
      "question": "Ile wsporników montażowych?",
      "type": "number",
      "field_name": "qty_brackets",
      "min": 1,
      "max": 20,
      "default": 4,
      "priority": "high",
      "why": "Bezpośrednio wpływa na czas realizacji"
    },
    {
      "question": "Jakie normy spawalnicze mają obowiązywać?",
      "type": "multi",
      "field_name": "welding_standards",
      "options": ["ISO 9606 (Europa)", "EN 1090 (konstrukcje stalowe)", "AWS D1.1 (USA)"],
      "priority": "medium",
      "why": "Normy definiują wymagania jakościowe i certyfikację"
    },
    {
      "question": "Czy wymagana jest obróbka CNC?",
      "type": "yes_no",
      "field_name": "cnc_required",
      "default": true,
      "priority": "medium",
      "why": "Obróbka CNC zwiększa precyzję ale +10-20% czasu"
    }
  ]
}

# ═══════════════════════════════════════════════════════════
# PRZYKŁADY
# ═══════════════════════════════════════════════════════════

PRZYKŁAD 1 - PEŁNY OPIS (brak pytań):
OPIS: "Płyta montażowa 400x300x5mm ze stali S235JR, 4 otwory gwintowane M8, spawanie MAG wg ISO 9606"

ODPOWIEDŹ:
{
  "components": [
    {"name": "Płyta bazowa 400x300x5mm", "layout_h": 1.0, "detail_h": 3.5, "doc_h": 1.5}
  ],
  "missing_info": [],  // PUSTE - wszystko jasne
  "suggestions": [
    {
      "type": "improvement",
      "title": "Dodaj fazowanie krawędzi",
      "description": "Fazowanie 1x45° zwiększy bezpieczeństwo i ułatwi montaż. +0.5h obróbki.",
      "priority": "low",
      "impact": {"hours_delta": 0.5}
    }
  ]
}

PRZYKŁAD 2 - BRAK SZCZEGÓŁÓW (dużo pytań):
OPIS: "Rama spawana z wspornikami"

ODPOWIEDŹ:
{
  "components": [
    {"name": "Rama główna (szacunkowo)", "layout_h": 3.0, "detail_h": 10.0, "doc_h": 5.0}
  ],
  "missing_info": [
    {
      "question": "Jaki materiał ramy?",
      "type": "choice",
      "field_name": "material_frame",
      "options": ["S235JR", "S355J2", "Aluminium"],
      "priority": "high",
      "why": "Materiał wpływa na czas obróbki i wagę konstrukcji"
    },
    {
      "question": "Ile wsporników?",
      "type": "number",
      "field_name": "qty_brackets",
      "min": 1, "max": 20, "default": 4,
      "priority": "high",
      "why": "Bezpośrednio wpływa na czas"
    },
    {
      "question": "Wymiary ramy (długość x szerokość x wysokość)?",
      "type": "text",
      "field_name": "frame_dimensions",
      "priority": "high",
      "why": "Duże wymiary (>5m) zwiększają złożoność +25%"
    },
    {
      "question": "Norma spawalnicza?",
      "type": "multi",
      "field_name": "welding_standards",
      "options": ["ISO 9606", "EN 1090", "AWS D1.1"],
      "priority": "medium",
      "why": "Normy definiują wymagania jakościowe"
    }
  ],
  "suggestions": [
    {
      "type": "warning",
      "title": "Uzupełnij opis dla dokładniejszej wyceny",
      "description": "Brakuje kluczowych informacji. Odpowiedz na pytania powyżej dla zwiększenia dokładności estymacji z ±40% do ±10%.",
      "priority": "high"
    }
  ]
}

WAŻNE: 
- Zwróć WYŁĄCZNIE JSON bez tekstu
- Każdy komponent z opisu = osobna pozycja
- ZAWSZE generuj "missing_info" jeśli czegoś brakuje
- ZAWSZE generuj min. 1 "suggestion"
- Godziny MUSZĄ odpowiadać złożoności
"""

# === HTTP Session z retry (stabilniejsze zapytania) ===
_session = None
def get_session():
    global _session
    if _session is None:
        s = requests.Session()
        retries = Retry(total=3, backoff_factor=0.5, status_forcelist=[429, 500, 502, 503, 504])
        s.mount('http://', HTTPAdapter(max_retries=retries))
        s.mount('https://', HTTPAdapter(max_retries=retries))
        _session = s
    return _session

# === NORMALIZACJA NAZW (canonical key) ===
def canonicalize_name(name: str) -> str:
    """Normalizuje nazwę komponentu do porównań i uczenia (z aliasami PL/DE/EN)."""
    if not name:
        return ""
    n = name.lower()
    # Usuń wymiary i liczby
    n = re.sub(r'\b\d+[.,]?\d*\s*(mm|cm|m|kg|t|ton|szt|pcs|inch|")\b', ' ', n)
    n = re.sub(r'\b\d+[.,]?\d*\b', ' ', n)
    # Tokenizacja i mapowanie aliasów
    tokens = re.split(r'[\s\-_.,;/]+', n)  # <-- NAPRAWIONE (escaped -)
    norm_tokens = []
    stoplist = {'i', 'a', 'the', 'and', 'or', 'der', 'die', 'das', 'und', 'ein', 'eine', 'of', 'for'}
    for tok in tokens:
        if not tok or tok in stoplist:
            continue
        mapped = COMPONENT_ALIASES.get(tok)
        if not mapped:
            for alias, canonical in COMPONENT_ALIASES.items():
                if alias in tok:
                    mapped = canonical
                    break
        norm_tokens.append(mapped or tok)
    seen, out = set(), []
    for t in norm_tokens:
        if t not in seen:
            seen.add(t)
            out.append(t)
    return ' '.join(out).strip()
    
# === Helpery integracji z doc-converterem i diagnostyki ===
def safe_json_loads(data_bytes_or_str):
    """Wczytuje JSON, toleruje code fences oraz bytes."""
    try:
        if isinstance(data_bytes_or_str, (bytes, bytearray)):
            s = data_bytes_or_str.decode("utf-8", errors="ignore")
        else:
            s = str(data_bytes_or_str)
        s = s.strip()
        if s.startswith("```json"):
            s = s[7:]
        if s.startswith("```"):
            s = s[3:]
        if s.endswith("```"):
            s = s[:-3]
        return json.loads(s)
    except Exception:
        return {}

def parse_components_from_docconv_json(obj: dict) -> list:
    """
    Wyciąga komponenty z JSON (różne schematy):
    - {"components":[{"name","hours_3d_layout","hours_3d_detail","hours_2d"}]}
    - {"components":[{"name","layout_h","detail_h","doc_h"}]}
    - {"components":[{"name","hours"}]} -> rozkłada na 30/50/20
    """
    if not isinstance(obj, dict):
        return []
    comps = obj.get("components") or []
    out = []
    for c in comps:
        if not isinstance(c, dict):
            continue
        name = c.get("name") or c.get("title") or c.get("component") or ""
        if not name:
            continue
        l = c.get("hours_3d_layout", c.get("layout_h", 0.0)) or 0.0
        d = c.get("hours_3d_detail", c.get("detail_h", 0.0)) or 0.0
        doc = c.get("hours_2d", c.get("doc_h", 0.0)) or 0.0
        if (l + d + doc) == 0 and c.get("hours"):
            tot = float(c.get("hours") or 0.0)
            l, d, doc = tot * 0.3, tot * 0.5, tot * 0.2
        item = {
            "name": name,
            "hours_3d_layout": float(l),
            "hours_3d_detail": float(d),
            "hours_2d": float(doc),
            "hours": float(l) + float(d) + float(doc),
            "is_summary": False,
            "comment": c.get("comment", "")
        }
        out.append(item)
    return out

def merge_components(base: list, extra: list) -> list:
    """Scala dwie listy komponentów, deduplikuje po canonicalize_name i sumuje godziny."""
    idx = {}
    out = []
    def key_of(c): return canonicalize_name(c.get("name",""))
    for c in base or []:
        k = key_of(c)
        if not k:
            out.append(c)
            continue
        idx[k] = dict(c)
    for c in extra or []:
        k = key_of(c)
        if not k:
            out.append(c)
            continue
        if k in idx:
            a = idx[k]
            a["hours_3d_layout"] = a.get("hours_3d_layout",0)+c.get("hours_3d_layout",0)
            a["hours_3d_detail"] = a.get("hours_3d_detail",0)+c.get("hours_3d_detail",0)
            a["hours_2d"] = a.get("hours_2d",0)+c.get("hours_2d",0)
            a["hours"] = a["hours_3d_layout"]+a["hours_3d_detail"]+a["hours_2d"]
        else:
            idx[k] = dict(c)
    # kolejnosc bazowych + nowe
    names_seen = set()
    for c in (base or []):
        k = key_of(c)
        if k and k in idx and k not in names_seen:
            out.append(idx[k]); names_seen.add(k)
        elif not k:
            out.append(c)
    for k, c in idx.items():
        if k not in names_seen:
            out.append(c)
    return out

def detect_embed_dim(model: str = EMBED_MODEL) -> int:
    """Zwraca długość wektora dla modelu embeddingowego (diagnostyka)."""
    try:
        v = get_embedding_ollama("embed-dim-probe", model=model)
        return len(v) if isinstance(v, list) else 0
    except Exception:
        return 0

# === WEKTORY / EMBEDDINGS ===
def to_pgvector(vec):
    if not vec:
        return None
    return '[' + ','.join(f'{float(x):.6f}' for x in vec) + ']'

@st.cache_data(ttl=86400, show_spinner=False)
def get_embedding_ollama(text: str, model: str = EMBED_MODEL) -> list:
    try:
        s = get_session()
        r = s.post(f"{OLLAMA_URL}/api/embeddings", json={"model": model, "prompt": text}, timeout=30)
        r.raise_for_status()
        return r.json().get("embedding", [])
    except Exception as e:
        logger.error(f"Embeddings error: {e}")
        return []

def ensure_project_embedding(cur, project_id: int, description: str):
    if not description or len(description.strip()) < 10:
        return
    emb = get_embedding_ollama(description)
    if emb and len(emb) == EMBED_DIM:
        try:
            cur.execute("UPDATE projects SET description_embedding = %s::vector WHERE id=%s",
                        (to_pgvector(emb), project_id))
        except Exception as e:
            logger.warning(f"Embedding failed for project {project_id}: {e}")

def ensure_pattern_embedding(cur, pattern_key: str, dept: str, text_for_embed: str):
    if not text_for_embed or len(text_for_embed.strip()) < 3:
        return
    emb = get_embedding_ollama(text_for_embed)
    if emb and len(emb) == EMBED_DIM:
        try:
            cur.execute("""
            UPDATE component_patterns
            SET name_embedding = %s::vector
            WHERE pattern_key=%s AND department=%s
            """, (to_pgvector(emb), pattern_key, dept))
        except Exception as e:
            logger.warning(f"Embedding failed for pattern {pattern_key}: {e}")

# === MODELE OLLAMA ===
@lru_cache(maxsize=1)
def list_local_models():
    try:
        r = requests.get(f"{OLLAMA_URL}/api/tags", timeout=5)
        if r.ok:
            return [m.get("name", "") for m in r.json().get("models", [])]
    except Exception:
        pass
    return []

def model_available(prefix: str) -> bool:
    return any(m.startswith(prefix) for m in list_local_models())

# === ZAPYTANIA DO OLLAMA (tekst + vision) ===
@st.cache_data(ttl=3600, show_spinner=False)
def query_ollama_cached(_payload_str: str) -> str:
    payload = json.loads(_payload_str)
    try:
        s = get_session()
        r = s.post(f"{OLLAMA_URL}/api/generate", json=payload, timeout=90)
        r.raise_for_status()
        return r.json().get('response', 'Brak odpowiedzi.')
    except Exception as e:
        logger.error(f"Błąd AI: {e}")
        return f"Błąd Ollama: {e}"

def query_ollama(prompt: str, model: str = "llama3:latest", images_b64=None, format_json=False) -> str:
    payload = {"model": model, "prompt": prompt, "stream": False}
    if images_b64:
        payload["images"] = images_b64
    if format_json:
        payload["format"] = "json"
    return query_ollama_cached(json.dumps(payload))

def encode_image_b64(file, max_px=1280, quality=85):
    """Kompresja obrazu do JPEG i zwrot Base64 (dla modeli Vision w Ollama)."""
    try:
        im = Image.open(file).convert("RGB")
        im.thumbnail((max_px, max_px))
        buf = io.BytesIO()
        im.save(buf, format="JPEG", quality=quality, optimize=True)
        return base64.b64encode(buf.getvalue()).decode("utf-8")
    except Exception as e:
        logger.warning(f"Błąd kompresji: {e}")
        return base64.b64encode(file.getvalue()).decode("utf-8")
# === CAD Estimator Pro — main.py (Part 2/4) ==================================
# Parsers (Excel/PDF), AI parsing, kategoryzacja, timeline, eksport XLSX,
# DB połączenia i podstawowe zapytania
# ============================================================================

# === PARSERY SUB-KOMPONENTÓW Z KOMENTARZY ===
def parse_subcomponents_from_comment(comment):
    """
    Ulepszony parser komentarzy:
    - obsługuje wpisy z i bez ilości,
    - toleruje myślnik po liczbie (np. '2x - docisk'),
    - odfiltrowuje ewidentne wymiary/jednostki.
    """
    if not comment or not isinstance(comment, str):
        return []

    def clean_name(s: str) -> str:
        s = re.sub(r'^\s*[-–—]\s*', '', s.strip())
        return s

    subcomponents = []
    qty_re = re.compile(
        r'(\d+)\s*(?:x|szt\.?|sztuk|pcs)?\s*[-–—]?\s*([^,;\n]+?)(?=[,;\n]|$)',
        re.IGNORECASE
    )

    consumed_spans = []
    for m in qty_re.finditer(comment):
        try:
            qty = int(m.group(1))
            name = clean_name(m.group(2))
            if len(name) >= 3 and not re.match(r'^\d+\s*(mm|cm|m|kg|ton|h)$', name, re.IGNORECASE):
                subcomponents.append({'quantity': qty, 'name': name})
                consumed_spans.append(m.span())
        except Exception:
            continue

    if consumed_spans:
        remainder_parts = []
        last = 0
        for a, b in consumed_spans:
            remainder_parts.append(comment[last:a])
            last = b
        remainder_parts.append(comment[last:])
        remainder = ';'.join(remainder_parts)
    else:
        remainder = comment

    for part in re.split(r'[;,]', remainder):
        name = clean_name(part)
        if not name or len(name) < 3:
            continue
        if re.match(r'^\d+(\.\d+)?\s*(mm|cm|m|kg|ton|h)?$', name, re.IGNORECASE):
            continue
        if qty_re.search(name):
            continue
        subcomponents.append({'quantity': 1, 'name': name})

    logger.info(f"Parsed {len(subcomponents)} subcomponents from: {comment[:80]}...")
    return subcomponents

# === PARSER ODPOWIEDZI AI (JSON → komponenty) ===
def parse_ai_response(text: str, components_from_excel=None):
    """Priorytet JSON, fallback regex, domknięcia sum i normalizacja risks/adjustments."""
    warnings = []
    parsed_components = []
    total_layout = total_detail = total_2d = 0.0
    data = {}

    if not text:
        warnings.append("Brak odpowiedzi od AI")
        return {
            "total_hours": 0.0, "total_layout": 0.0, "total_detail": 0.0, "total_2d": 0.0,
            "components": components_from_excel or [], "raw_text": "", "warnings": warnings,
            "analysis": {}, "missing_info": [], "phases": {},
            "risks_detailed": [], "recommendations": [], "ai_adjustments": []
        }

    clean = text.strip()
    if clean.startswith("```json"):
        clean = clean[7:]
    if clean.startswith("```"):
        clean = clean[3:]
    if clean.endswith("```"):
        clean = clean[:-3]

    try:
        data = json.loads(clean)

        # Normalizacja ryzyk
        risks = []
        for r in data.get("risks", []):
            if isinstance(r, str):
                risks.append({"risk": r, "impact": "nieznany", "mitigation": "Do określenia"})
            else:
                risks.append({
                    "risk": r.get("risk", "Nieznane ryzyko"),
                    "impact": r.get("impact", "nieznany"),
                    "mitigation": r.get("mitigation", "Brak")
                })
        data["risks"] = risks

        # Komponenty
        for c in data.get("components", []):
            item = {
                "name": c.get("name", "bez nazwy"),
                "hours_3d_layout": float(c.get("layout_h", 0) or 0),
                "hours_3d_detail": float(c.get("detail_h", 0) or 0),
                "hours_2d": float(c.get("doc_h", 0) or 0),
            }
            item["hours"] = item["hours_3d_layout"] + item["hours_3d_detail"] + item["hours_2d"]
            item["is_summary"] = False
            parsed_components.append(item)

        sums = data.get("sums", {})
        total_layout = float(sums.get("layout", 0) or sum(x["hours_3d_layout"] for x in parsed_components))
        total_detail = float(sums.get("detail", 0) or sum(x["hours_3d_detail"] for x in parsed_components))
        total_2d = float(sums.get("doc", 0) or sum(x["hours_2d"] for x in parsed_components))

        # Adjustments (AI)
        ai_adj = []
        for ad in data.get("adjustments", []):
            parent = ad.get("parent")
            adds = ad.get("adds", [])
            norm_adds = []
            for a in adds:
                norm_adds.append({
                    "name": a.get("name", "sub"),
                    "qty": int(a.get("qty", 1) or 1),
                    "layout_add": float(a.get("layout_add", 0) or 0),
                    "detail_add": float(a.get("detail_add", 0) or 0),
                    "doc_add": float(a.get("doc_add", 0) or 0),
                    "reason": a.get("reason", "")
                })
            ai_adj.append({"parent": parent, "adds": norm_adds})
        data["ai_adjustments"] = ai_adj

    except json.JSONDecodeError:
        warnings.append("Fallback do regex (AI nie zwrócił poprawnego JSON)")
        pattern = r"-\s*([^\n]+?)\s+Layout:\s*(\d+[.,]?\d*)\s*h?,?\s*Detail:\s*(\d+[.,]?\d*)\s*h?,?\s*2D:\s*(\d+[.,]?\d*)\s*h?"
        for m in re.finditer(pattern, text, re.IGNORECASE):
            try:
                parsed_components.append({
                    "name": m.group(1).strip(),
                    "hours_3d_layout": float(m.group(2).replace(',', '.')),
                    "hours_3d_detail": float(m.group(3).replace(',', '.')),
                    "hours_2d": float(m.group(4).replace(',', '.')),
                    "hours": sum(float(m.group(i).replace(',', '.')) for i in [2,3,4]),
                    "is_summary": False
                })
            except Exception:
                pass
        data["ai_adjustments"] = []

    # fallback do Excela
    excel_parts = [c for c in (components_from_excel or []) if not c.get('is_summary', False)]
    if not parsed_components and excel_parts:
        warnings.append("Użyto danych z Excel - AI nie zwróciło komponentów")
        parsed_components = excel_parts
    elif parsed_components and excel_parts and len(parsed_components) < len(excel_parts) * 0.5:
        warnings.append(f"AI zwróciło tylko {len(parsed_components)} z {len(excel_parts)} komponentów - użyto danych z Excel")
        parsed_components = excel_parts

    if total_layout == 0 and parsed_components:
        total_layout = sum(c.get('hours_3d_layout', 0) for c in parsed_components)
    if total_detail == 0 and parsed_components:
        total_detail = sum(c.get('hours_3d_detail', 0) for c in parsed_components)
    if total_2d == 0 and parsed_components:
        total_2d = sum(c.get('hours_2d', 0) for c in parsed_components)

    # ═══════════════════════════════════════════════════════════
    # NOWE: Parsowanie suggestions i missing_info
    # ═══════════════════════════════════════════════════════════
    
    suggestions = []
    for s in data.get("suggestions", []):
        if isinstance(s, dict):
            suggestions.append({
                "type": s.get("type", "other"),
                "title": s.get("title", "Sugestia"),
                "description": s.get("description", ""),
                "priority": s.get("priority", "medium"),
                "impact": s.get("impact", {}),
                "components_to_add": s.get("components_to_add", []),
                "components_to_remove": s.get("components_to_remove", [])
            })
    
    missing_info = []
    for m in data.get("missing_info", []):
        if isinstance(m, dict):
            missing_info.append({
                "question": m.get("question", "Pytanie"),
                "type": m.get("type", "text"),
                "field_name": m.get("field_name", f"field_{len(missing_info)}"),
                "options": m.get("options", []),
                "default": m.get("default"),
                "min": m.get("min"),
                "max": m.get("max"),
                "priority": m.get("priority", "medium"),
                "why": m.get("why", "")
            })

    return {
        "total_hours": max(0.0, total_layout + total_detail + total_2d),
        "total_layout": total_layout,
        "total_detail": total_detail,
        "total_2d": total_2d,
        "components": parsed_components,
        "raw_text": text,
        "warnings": warnings,
        "analysis": data.get("analysis", {}),
        "missing_info": missing_info,  # ⬅️ NOWE
        "phases": data.get("phases", {}),
        "risks_detailed": data.get("risks", []),
        "recommendations": data.get("recommendations", []),
        "ai_adjustments": data.get("ai_adjustments", []),
        "suggestions": suggestions  # ⬅️ NOWE
    }
    
    

# === PARSERY EXCEL (z/bez komentarzy) ===
def parse_cad_project_structured(file_stream):
    """Parser Excel z hierarchią, komentarzami i godzinami (bez openpyxl-comments)."""
    result = {'components': [], 'multipliers': {}, 'totals': {}, 'statistics': {}}
    df = pd.read_excel(file_stream, header=None)

    # Kolumny
    COL_POS, COL_DESC, COL_COMMENT = 0, 1, 2
    COL_STD_PARTS, COL_SPEC_PARTS = 3, 4
    COL_HOURS_LAYOUT, COL_HOURS_DETAIL, COL_HOURS_DOC = 7, 9, 11

    # Multipliers
    try:
        result['multipliers']['layout'] = float(df.iloc[9, COL_HOURS_LAYOUT]) if pd.notna(df.iloc[9, COL_HOURS_LAYOUT]) else 1.0
        result['multipliers']['detail'] = float(df.iloc[9, COL_HOURS_DETAIL]) if pd.notna(df.iloc[9, COL_HOURS_DETAIL]) else 1.0
        result['multipliers']['documentation'] = float(df.iloc[9, COL_HOURS_DOC]) if pd.notna(df.iloc[9, COL_HOURS_DOC]) else 1.0
    except Exception:
        result['multipliers'] = {'layout': 1.0, 'detail': 1.0, 'documentation': 1.0}

    data_start_row = 11
    for row_idx in range(data_start_row, df.shape[0]):
        try:
            pos = str(df.iloc[row_idx, COL_POS]).strip() if pd.notna(df.iloc[row_idx, COL_POS]) else ""
            name = str(df.iloc[row_idx, COL_DESC]).strip() if pd.notna(df.iloc[row_idx, COL_DESC]) else ""
            if not pos or pos in ['nan', 'None', '']: 
                continue
            if not name or name in ['nan', 'None']: 
                name = f"[Pozycja {pos}]"
            comment = str(df.iloc[row_idx, COL_COMMENT]).strip() if pd.notna(df.iloc[row_idx, COL_COMMENT]) else ""

            hours_layout = float(df.iloc[row_idx, COL_HOURS_LAYOUT]) if pd.notna(df.iloc[row_idx, COL_HOURS_LAYOUT]) else 0.0
            hours_detail = float(df.iloc[row_idx, COL_HOURS_DETAIL]) if pd.notna(df.iloc[row_idx, COL_HOURS_DETAIL]) else 0.0
            hours_doc = float(df.iloc[row_idx, COL_HOURS_DOC]) if pd.notna(df.iloc[row_idx, COL_HOURS_DOC]) else 0.0

            is_summary = bool(re.match(r'^\d+,0$', pos) or pos.isdigit())
            subcomponents = parse_subcomponents_from_comment(comment)

            component = {
                'id': pos, 'name': name, 'comment': comment,
                'type': 'assembly' if is_summary else 'part',
                'level': pos.count(','),
                'parts': {
                    'standard': int(float(df.iloc[row_idx, COL_STD_PARTS])) if pd.notna(df.iloc[row_idx, COL_STD_PARTS]) else 0,
                    'special': int(float(df.iloc[row_idx, COL_SPEC_PARTS])) if pd.notna(df.iloc[row_idx, COL_SPEC_PARTS]) else 0
                },
                'hours_3d_layout': hours_layout,
                'hours_3d_detail': hours_detail,
                'hours_2d': hours_doc,
                'hours': hours_layout + hours_detail + hours_doc,
                'is_summary': is_summary,
                'subcomponents': subcomponents
            }
            result['components'].append(component)
        except Exception as e:
            logger.warning(f"Błąd wiersz {row_idx + 1}: {e}")
            continue

    parts_only = [
        c for c in result['components']
        if not c.get('is_summary', False) and c.get('hours', 0) > 0 and c.get('name') not in ['[part]', '[assembly]', '', ' ']
    ]
    result['totals']['layout'] = sum(c['hours_3d_layout'] for c in parts_only)
    result['totals']['detail'] = sum(c['hours_3d_detail'] for c in parts_only)
    result['totals']['documentation'] = sum(c['hours_2d'] for c in parts_only)
    result['totals']['total'] = sum(c['hours'] for c in parts_only)
    result['statistics']['parts_count'] = len(parts_only)
    result['statistics']['assemblies_count'] = sum(1 for c in result['components'] if c.get('is_summary', False))
    return result

def parse_cad_project_structured_with_xlsx_comments(file_like):
    """
    Parser Excel z odczytem:
    - wartości (pandas),
    - komentarzy/note komórek (openpyxl),
    - łączeniem komentarzy z całego wiersza.
    Zwraca strukturę identyczną jak parse_cad_project_structured.
    """
    # Wczytaj bytes i utwórz dwa niezależne strumienie
    if hasattr(file_like, "read"):
        content = file_like.read()
    elif isinstance(file_like, bytes):
        content = file_like
    else:
        file_like.seek(0)
        content = file_like.read()

    bio_pd = BytesIO(content)
    bio_xl = BytesIO(content)

    # 1) Dane tabelaryczne
    df = pd.read_excel(bio_pd, header=None)

    # 2) Komentarze komórek
    comments_map = {}
    try:
        wb = load_workbook(bio_xl, data_only=True)
        ws = wb.active
        for r in ws.iter_rows():
            for cell in r:
                if cell.comment and cell.comment.text:
                    comments_map[(cell.row - 1, cell.column - 1)] = cell.comment.text.strip()
    except Exception as e:
        logger.info(f"Brak/nie udało się odczytać komentarzy z xlsx: {e}")

    result = {'components': [], 'multipliers': {}, 'totals': {}, 'statistics': {}}

    # Kolumny zgodne z parserem podstawowym
    COL_POS, COL_DESC, COL_COMMENT = 0, 1, 2
    COL_STD_PARTS, COL_SPEC_PARTS = 3, 4
    COL_HOURS_LAYOUT, COL_HOURS_DETAIL, COL_HOURS_DOC = 7, 9, 11

    # Multipliers
    try:
        result['multipliers']['layout'] = float(df.iloc[9, COL_HOURS_LAYOUT]) if pd.notna(df.iloc[9, COL_HOURS_LAYOUT]) else 1.0
        result['multipliers']['detail'] = float(df.iloc[9, COL_HOURS_DETAIL]) if pd.notna(df.iloc[9, COL_HOURS_DETAIL]) else 1.0
        result['multipliers']['documentation'] = float(df.iloc[9, COL_HOURS_DOC]) if pd.notna(df.iloc[9, COL_HOURS_DOC]) else 1.0
    except Exception:
        result['multipliers'] = {'layout': 1.0, 'detail': 1.0, 'documentation': 1.0}

    data_start_row = 11
    for row_idx in range(data_start_row, df.shape[0]):
        try:
            pos = str(df.iloc[row_idx, COL_POS]).strip() if pd.notna(df.iloc[row_idx, COL_POS]) else ""
            name = str(df.iloc[row_idx, COL_DESC]).strip() if pd.notna(df.iloc[row_idx, COL_DESC]) else ""
            if not pos or pos in ['nan', 'None', '']:
                continue
            if not name or name in ['nan', 'None']:
                name = f"[Pozycja {pos}]"

            # Tekst z kolumny "Komentarz"
            txt_comment_col = str(df.iloc[row_idx, COL_COMMENT]).strip() if pd.notna(df.iloc[row_idx, COL_COMMENT]) else ""

            # Zbierz komentarze z całego wiersza
            row_comments = []
            try:
                max_cols = max(c for (_, c) in comments_map.keys() if _ == row_idx) + 1 if comments_map else df.shape[1]
            except ValueError:
                max_cols = df.shape[1]

            for col in range(max_cols):
                txt = comments_map.get((row_idx, col))
                if txt:
                    row_comments.append(txt)

            joined_cell_comments = "; ".join([t for t in row_comments if t])
            comment = "; ".join([t for t in [txt_comment_col, joined_cell_comments] if t])

            hours_layout = float(df.iloc[row_idx, COL_HOURS_LAYOUT]) if pd.notna(df.iloc[row_idx, COL_HOURS_LAYOUT]) else 0.0
            hours_detail = float(df.iloc[row_idx, COL_HOURS_DETAIL]) if pd.notna(df.iloc[row_idx, COL_HOURS_DETAIL]) else 0.0
            hours_doc = float(df.iloc[row_idx, COL_HOURS_DOC]) if pd.notna(df.iloc[row_idx, COL_HOURS_DOC]) else 0.0

            is_summary = bool(re.match(r'^\d+,0$', pos) or pos.isdigit())
            subcomponents = parse_subcomponents_from_comment(comment)

            component = {
                'id': pos, 'name': name, 'comment': comment,
                'type': 'assembly' if is_summary else 'part',
                'level': pos.count(','),
                'parts': {
                    'standard': int(float(df.iloc[row_idx, COL_STD_PARTS])) if pd.notna(df.iloc[row_idx, COL_STD_PARTS]) else 0,
                    'special': int(float(df.iloc[row_idx, COL_SPEC_PARTS])) if pd.notna(df.iloc[row_idx, COL_SPEC_PARTS]) else 0
                },
                'hours_3d_layout': hours_layout,
                'hours_3d_detail': hours_detail,
                'hours_2d': hours_doc,
                'hours': hours_layout + hours_detail + hours_doc,
                'is_summary': is_summary,
                'subcomponents': subcomponents
            }
            result['components'].append(component)
        except Exception as e:
            logger.warning(f"Błąd wiersz {row_idx + 1}: {e}")
            continue

    parts_only = [
        c for c in result['components']
        if not c.get('is_summary', False) and c.get('hours', 0) > 0 and c.get('name') not in ['[part]', '[assembly]', '', ' ']
    ]
    result['totals']['layout'] = sum(c['hours_3d_layout'] for c in parts_only)
    result['totals']['detail'] = sum(c['hours_3d_detail'] for c in parts_only)
    result['totals']['documentation'] = sum(c['hours_2d'] for c in parts_only)
    result['totals']['total'] = sum(c['hours'] for c in parts_only)
    result['statistics']['parts_count'] = len(parts_only)
    result['statistics']['assemblies_count'] = sum(1 for c in result['components'] if c.get('is_summary', False))
    return result

def process_excel(file):
    """Wczytuje plik Excel (bytes) i próbuje parser z komentarzami, a potem fallback."""
    try:
        content = file.read()
        bio1 = BytesIO(content)
        bio2 = BytesIO(content)

        try:
            result = parse_cad_project_structured_with_xlsx_comments(bio1)
            used_parser = "xlsx+comments"
        except Exception as e:
            logger.info(f"Parser z komentarzami nieudany, fallback: {e}")
            result = parse_cad_project_structured(bio2)
            used_parser = "basic"

        if result and result.get('components'):
            parts_only = [c for c in result['components'] if not c.get('is_summary', False)]
            st.success(f"✅ {len(parts_only)} komponentów: Layout {result['totals']['layout']:.1f}h + Detail {result['totals']['detail']:.1f}h + 2D {result['totals']['documentation']:.1f}h (parser: {used_parser})")
            if result.get('multipliers'):
                st.info(f"Współczynniki: Layout={result['multipliers']['layout']}, Detail={result['multipliers']['detail']}, Doc={result['multipliers']['documentation']}")
                st.session_state["excel_multipliers"] = result['multipliers']
            return result['components']
        else:
            st.warning("Brak komponentów w pliku")
            return []
    except Exception as e:
        st.error(f"Błąd parsowania: {e}")
        logger.exception("Błąd parsowania Excel")
        return []

# === PDF ===
def extract_text_from_pdf(pdf_file):
    """Lekki ekstraktor tekstu PDF (PyPDF2)."""
    try:
        reader = PdfReader(pdf_file)
        text = ""
        max_pages = 200
        for i, page in enumerate(reader.pages):
            if i >= max_pages:
                text += f"\n[... {len(reader.pages)} stron, przetworzono {max_pages} ...]"
                break
            text += (page.extract_text() or "") + "\n"
        logger.info(f"PDF: {len(text)} znaków")
        return text
    except Exception as e:
        logger.error(f"Błąd PDF: {e}")
        return f"[Błąd PDF: {e}]"

# === KATEGORYZACJA & TIMELINE ===
def categorize_component(name: str) -> str:
    categories = {
        "analiza": ["przegląd", "analiza", "normy"],
        "modelowanie": ["modelowanie", "3d", "konstrukcja"],
        "obliczenia": ["obliczenia", "mes", "fem"],
        "rysunki": ["rysunek", "dokumentacja", "bom"],
        "spawanie": ["spawanie", "weld"],
        "automatyka": ["plc", "robot"]
    }
    n = name.lower()
    for cat, keys in categories.items():
        if any(k in n for k in keys):
            return cat
    return "inne"

def show_project_timeline(components):
    """Prosty wykres sekwencyjny (plotly) – suma godzin L/D/2D per komponent."""
    if not components:
        st.info("Brak komponentów do wyświetlenia")
        return
    parts = [c for c in components if not c.get('is_summary', False) and c.get('hours', 0) > 0]
    if not parts:
        st.info("Brak komponentów z godzinami")
        return
    timeline_data = []
    cumulative = 0.0
    for comp in parts:
        hours = float(comp.get('hours', 0) or 0)
        timeline_data.append({
            'Task': comp['name'][:30] + "..." if len(comp['name']) > 30 else comp['name'],
            'Start': cumulative,
            'Finish': cumulative + hours,
            'Hours': hours
        })
        cumulative += hours
    df = pd.DataFrame(timeline_data)
    fig = px.bar(df, x='Hours', y='Task', orientation='h',
                 title="Harmonogram realizacji (sekwencyjnie)",
                 labels={'Hours': 'Godziny', 'Task': 'Komponent'})
    fig.update_layout(yaxis={'categoryorder': 'total ascending'})
    st.plotly_chart(fig, use_container_width=True)

# === EKSPORT DO EXCEL (dynamiczny engine) ===
def export_quotation_to_excel(project_data):
    """
    Eksport bez twardej zależności:
    - preferuj xlsxwriter (jeśli zainstalowany),
    - fallback na openpyxl,
    - jeśli brak obu — podnieś czytelny wyjątek.
    """
    output = BytesIO()
    engine = None
    try:
        import xlsxwriter  # noqa
        engine = "xlsxwriter"
    except Exception:
        try:
            import openpyxl  # noqa
            engine = "openpyxl"
        except Exception:
            raise RuntimeError("Brak silnika do zapisu XLSX. Zainstaluj: xlsxwriter lub openpyxl.")

    with pd.ExcelWriter(output, engine=engine) as writer:
        components = [c for c in project_data.get('components', []) if not c.get('is_summary', False)]
        df_components = pd.DataFrame(components)
        if not df_components.empty:
            df_components.to_excel(writer, sheet_name='Wycena', index=False)
        else:
            pd.DataFrame([{"info": "Brak"}]).to_excel(writer, sheet_name='Wycena', index=False)
        summary = pd.DataFrame({
            'Parametr': ['Nazwa', 'Klient', 'Dział', 'Suma', 'Data'],
            'Wartość': [
                project_data.get('name', ''), project_data.get('client', ''),
                project_data.get('department', ''), f"{project_data.get('total_hours', 0):.1f}",
                datetime.now().strftime('%Y-%m-%d')
            ]
        })
        summary.to_excel(writer, sheet_name='Podsumowanie', index=False)
    output.seek(0)
    return output.getvalue()

# === DB: POŁĄCZENIA I INICJALIZACJA ===
@contextmanager
def get_db_connection():
    conn = None
    try:
        conn = psycopg2.connect(
            host=DB_HOST, database=DB_NAME, user=DB_USER, password=DB_PASSWORD, port=5432
        )
        yield conn
    except psycopg2.OperationalError as e:
        logger.error(f"Błąd połączenia: {e}")
        st.error("Błąd połączenia z bazą.")
        st.stop()
    except Exception as e:
        logger.error(f"Błąd: {e}")
        if conn:
            conn.rollback()
        raise
    finally:
        if conn:
            conn.close()

@st.cache_resource
def init_db():
    """Tworzy rozszerzenia, tabele, indeksy oraz nowe kolumny (idempotentnie)."""
    try:
        with get_db_connection() as conn, conn.cursor() as cur:
            # Extension
            cur.execute("CREATE EXTENSION IF NOT EXISTS vector;")

            # Tabele
            cur.execute('''
            CREATE TABLE IF NOT EXISTS projects (
                id SERIAL PRIMARY KEY,
                name VARCHAR(255) NOT NULL,
                client VARCHAR(255),
                department VARCHAR(10),
                cad_system VARCHAR(50),
                components JSONB,
                estimated_hours_3d_layout FLOAT DEFAULT 0,
                estimated_hours_3d_detail FLOAT DEFAULT 0,
                estimated_hours_2d FLOAT DEFAULT 0,
                estimated_hours FLOAT,
                actual_hours FLOAT,
                complexity_score INTEGER,
                accuracy FLOAT,
                description TEXT,
                ai_analysis TEXT,
                created_at TIMESTAMP DEFAULT NOW(),
                updated_at TIMESTAMP DEFAULT NOW(),
                description_embedding vector(%s),
                is_historical BOOLEAN DEFAULT FALSE,
                estimation_mode VARCHAR(30) DEFAULT 'ai',
                totals_source VARCHAR(30) DEFAULT 'ai',
                locked_totals BOOLEAN DEFAULT FALSE
            )
            ''', (EMBED_DIM,))

            cur.execute('''
            CREATE TABLE IF NOT EXISTS component_patterns (
                id SERIAL PRIMARY KEY,
                name VARCHAR(255) NOT NULL,
                department VARCHAR(10),
                avg_hours_3d_layout FLOAT DEFAULT 0,
                avg_hours_3d_detail FLOAT DEFAULT 0,
                avg_hours_2d FLOAT DEFAULT 0,
                avg_hours_total FLOAT DEFAULT 0,
                proportion_layout FLOAT DEFAULT 0.33,
                proportion_detail FLOAT DEFAULT 0.33,
                proportion_doc FLOAT DEFAULT 0.33,
                std_dev_hours FLOAT DEFAULT 0,
                occurrences INTEGER DEFAULT 0,
                min_hours FLOAT,
                max_hours FLOAT,
                typical_complexity FLOAT DEFAULT 1.0,
                cad_systems JSONB,
                last_updated TIMESTAMP DEFAULT NOW(),
                pattern_key TEXT,
                name_embedding vector(%s),
                m2_layout DOUBLE PRECISION DEFAULT 0,
                m2_detail DOUBLE PRECISION DEFAULT 0,
                m2_doc DOUBLE PRECISION DEFAULT 0,
                m2_total DOUBLE PRECISION DEFAULT 0,
                confidence DOUBLE PRECISION DEFAULT 0,
                source TEXT,
                last_actual_sample_at TIMESTAMP,
                UNIQUE(name, department)
            )
            ''', (EMBED_DIM,))

            cur.execute('''
            CREATE TABLE IF NOT EXISTS project_versions (
                id SERIAL PRIMARY KEY,
                project_id INTEGER REFERENCES projects(id) ON DELETE CASCADE,
                version VARCHAR(20) NOT NULL,
                components JSONB,
                estimated_hours FLOAT,
                estimated_hours_3d_layout FLOAT DEFAULT 0,
                estimated_hours_3d_detail FLOAT DEFAULT 0,
                estimated_hours_2d FLOAT DEFAULT 0,
                change_description TEXT,
                changed_by VARCHAR(100),
                is_approved BOOLEAN DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT NOW()
            )
            ''')

            cur.execute('''
            CREATE TABLE IF NOT EXISTS category_baselines (
                id SERIAL PRIMARY KEY,
                department VARCHAR(10),
                category TEXT,
                mean_layout DOUBLE PRECISION DEFAULT 0,
                mean_detail DOUBLE PRECISION DEFAULT 0,
                mean_doc DOUBLE PRECISION DEFAULT 0,
                m2_layout DOUBLE PRECISION DEFAULT 0,
                m2_detail DOUBLE PRECISION DEFAULT 0,
                m2_doc DOUBLE PRECISION DEFAULT 0,
                occurrences INTEGER DEFAULT 0,
                confidence DOUBLE PRECISION DEFAULT 0,
                last_updated TIMESTAMP DEFAULT NOW(),
                UNIQUE(department, category)
            )
            ''')

            cur.execute('''
            CREATE TABLE IF NOT EXISTS component_bundles (
                id SERIAL PRIMARY KEY,
                department VARCHAR(10),
                parent_key TEXT,
                parent_name TEXT,
                sub_key TEXT,
                sub_name TEXT,
                occurrences INTEGER DEFAULT 0,
                total_qty DOUBLE PRECISION DEFAULT 0,
                confidence DOUBLE PRECISION DEFAULT 0,
                last_updated TIMESTAMP DEFAULT NOW(),
                UNIQUE (department, parent_key, sub_key)
            )
            ''')

            # Indeksy
            cur.execute('CREATE INDEX IF NOT EXISTS idx_projects_created_at ON projects(created_at DESC)')
            cur.execute('CREATE INDEX IF NOT EXISTS idx_projects_department ON projects(department)')
            cur.execute('CREATE INDEX IF NOT EXISTS idx_patterns_department ON component_patterns(department)')
            cur.execute('CREATE INDEX IF NOT EXISTS idx_component_patterns_name ON component_patterns(name, department)')
            cur.execute('CREATE INDEX IF NOT EXISTS idx_versions_project ON project_versions(project_id, created_at DESC)')
            cur.execute("CREATE INDEX IF NOT EXISTS idx_projects_hist ON projects(is_historical)")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_projects_estimation_mode ON projects(estimation_mode)")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_projects_totals_source ON projects(totals_source)")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_projects_locked_totals ON projects(locked_totals)")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_projects_desc_embed_hnsw ON projects USING hnsw (description_embedding vector_l2_ops)")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_component_patterns_name_embed_hnsw ON component_patterns USING hnsw (name_embedding vector_l2_ops)")
            cur.execute('CREATE INDEX IF NOT EXISTS idx_bundles_dept_parent ON component_bundles(department, parent_key)')
            cur.execute('CREATE INDEX IF NOT EXISTS idx_bundles_dept_parent_occ ON component_bundles(department, parent_key, occurrences DESC)')

            conn.commit()
            logger.info("Baza zainicjalizowana + migracje/indeksy.")
            return True
    except Exception as e:
        logger.error(f"Błąd inicjalizacji: {e}")
        st.error(f"Błąd inicjalizacji: {e}")
        return False

# === OPERACJE NA WERSJACH I ZAPYTANIA PODOBIEŃSTWA ===
def save_project_version(conn, project_id, version, components, estimated_hours, layout_h, detail_h, doc_h, change_desc, changed_by):
    with conn.cursor() as cur:
        cur.execute("""
        INSERT INTO project_versions (
            project_id, version, components, estimated_hours,
            estimated_hours_3d_layout, estimated_hours_3d_detail, estimated_hours_2d,
            change_description, changed_by
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id
        """, (project_id, version, json.dumps(components, ensure_ascii=False),
              estimated_hours, layout_h, detail_h, doc_h, change_desc, changed_by))
        version_id = cur.fetchone()[0]
        conn.commit()
        return version_id

def get_project_versions(conn, project_id):
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute("""
        SELECT id, version, estimated_hours, estimated_hours_3d_layout,
               estimated_hours_3d_detail, estimated_hours_2d,
               change_description, changed_by, is_approved, created_at
        FROM project_versions WHERE project_id = %s ORDER BY created_at DESC
        """, (project_id,))
        return cur.fetchall()

def find_similar_projects(conn, description, department, limit=3):
    if not description:
        return []
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute("""
        SELECT id, name, client, estimated_hours, actual_hours, department
        FROM projects
        WHERE department = %s
        AND to_tsvector('simple', coalesce(name,'') || ' ' || coalesce(client,'') || ' ' || coalesce(description,'')) 
            @@ websearch_to_tsquery('simple', %s)
        ORDER BY created_at DESC LIMIT %s
        """, (department, description, limit))
        return cur.fetchall()

def find_similar_projects_semantic(conn, description, department, limit=5):
    if not description:
        return []
    emb = get_embedding_ollama(description)
    if not emb:
        return []
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute("""
        SELECT id, name, client, department, estimated_hours, actual_hours,
               1 - (description_embedding <-> %s::vector) AS similarity
        FROM projects
        WHERE department = %s AND description_embedding IS NOT NULL
        ORDER BY description_embedding <-> %s::vector
        LIMIT %s
        """, (to_pgvector(emb), department, to_pgvector(emb), limit))
        return cur.fetchall()

def find_similar_components(conn, name, department, limit=5):
    key = canonicalize_name(name)
    emb = get_embedding_ollama(key)
    if not emb:
        return []
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute("""
        SELECT name, avg_hours_total, avg_hours_3d_layout, avg_hours_3d_detail, avg_hours_2d,
               confidence, occurrences, 1 - (name_embedding <-> %s::vector) AS similarity
        FROM component_patterns
        WHERE department=%s AND name_embedding IS NOT NULL
        ORDER BY name_embedding <-> %s::vector
        LIMIT %s
        """, (to_pgvector(emb), department, to_pgvector(emb), limit))
        return cur.fetchall()
# === CAD Estimator Pro — main.py (Part 3/4) ==================================
# Uczenie (patterns/bundles), heurystyki, propozycje dodatków,
# Batch import oraz strona "Nowy projekt" z JSON/paste i Vision (llava/qwen2-vl)
# ============================================================================

# === STATYSTYKA (Welford) i dopasowanie kluczy ===
def _welford_step(mean, m2, n, x):
    """
    Algorytm Welforda - aktualizacja średniej i wariancji online.
    Zapobiega overflow i jest numerycznie stabilny.
    
    Outlier detection: jeśli n >= 5, odrzuca wartości > 2.5 std od średniej.
    """
    # Outlier detection (po zebraniu przynajmniej 5 próbek)
    if n and n >= 5:
        std = (m2 / max(n - 1, 1)) ** 0.5
        if mean and abs(x - mean) > 3.5 * std:
            logger.debug(f"   ⚠️ Outlier odrzucony: x={x:.2f}, mean={mean:.2f}, std={std:.2f}")
            return mean, m2, n  # Nie aktualizuj - outlier
    
    # Welford update
    n_new = (n or 0) + 1
    delta = x - (mean or 0)
    mean_new = (mean or 0) + delta / n_new
    delta2 = x - mean_new
    m2_new = (m2 or 0) + delta * delta2
    
    return mean_new, m2_new, n_new
def best_pattern_key(cur, dept: str, key: str, threshold: int = 88) -> str:
    """Jeśli pattern_key nie istnieje – dopasuj fuzzy do istniejących w danym dziale."""
    cur.execute("SELECT pattern_key FROM component_patterns WHERE pattern_key=%s AND department=%s", (key, dept))
    if cur.fetchone():
        return key
    cur.execute("SELECT DISTINCT pattern_key FROM component_patterns WHERE department=%s AND pattern_key IS NOT NULL", (dept,))
    keys = [r[0] for r in cur.fetchall()]
    if not keys:
        return key
    match, score, _ = process.extractOne(key, keys, scorer=fuzz.token_sort_ratio)
    return match if score >= threshold else key


def update_pattern_smart(cur, name, dept, layout_h, detail_h, doc_h, source='actual'):
    """
    Uczy wzorce z algorytmem Welford (running average + variance).
    Zapobiega tworzeniu duplikatów - aktualizuje istniejące wzorce.
    """
    try:
        # Walidacja
        if not name or not name.strip():
            return False
        if not dept:
            return False
        
        # Normalizuj klucz (z poprawionym regex)
        pattern_key = canonicalize_name(name)
        if not pattern_key:
            logger.warning(f"⚠️ Pominięto wzorzec '{name}': pusta pattern_key po normalizacji")
            return False
        
        # Oblicz total
        layout_h = float(layout_h or 0)
        detail_h = float(detail_h or 0)
        doc_h = float(doc_h or 0)
        total_h = layout_h + detail_h + doc_h
        
        logger.debug(f"📝 Uczę wzorzec: '{name[:40]}' → key='{pattern_key}', dept={dept}, total={total_h:.2f}h")
        
        # Sprawdź czy istnieje
        cur.execute("""
            SELECT occurrences, 
                   avg_hours_3d_layout, avg_hours_3d_detail, avg_hours_2d, avg_hours_total,
                   m2_layout, m2_detail, m2_doc, m2_total
            FROM component_patterns
            WHERE pattern_key = %s AND department = %s
        """, (pattern_key, dept))
        
        existing = cur.fetchone()
        
        if existing:
            # UPDATE z Welford
            old_occ, old_layout, old_detail, old_doc, old_total = existing[0:5]
            m2_layout, m2_detail, m2_doc, m2_total = existing[5:9]
            
            # Welford update
            new_occ = (old_occ or 0) + 1
            
            # Mean update: new_mean = old_mean + (new_value - old_mean) / new_count
            new_layout, m2_layout, _ = _welford_step(old_layout, m2_layout, old_occ, layout_h)
            new_detail, m2_detail, _ = _welford_step(old_detail, m2_detail, old_occ, detail_h)
            new_doc, m2_doc, _ = _welford_step(old_doc, m2_doc, old_occ, doc_h)
            new_total, m2_total, _ = _welford_step(old_total, m2_total, old_occ, total_h)
            
            # Confidence based on count and variance
            std_total = (m2_total / max(new_occ - 1, 1)) ** 0.5 if new_occ > 1 else 0.0
            confidence = min(1.0, new_occ / 10.0) * (1.0 / (1.0 + (std_total / (new_total or 1e-6))))
            
            cur.execute("""
                UPDATE component_patterns
                SET avg_hours_3d_layout = %s,
                    avg_hours_3d_detail = %s,
                    avg_hours_2d = %s,
                    avg_hours_total = %s,
                    m2_layout = %s,
                    m2_detail = %s,
                    m2_doc = %s,
                    m2_total = %s,
                    occurrences = %s,
                    confidence = %s,
                    source = %s,
                    last_updated = NOW(),
                    last_actual_sample_at = CASE WHEN %s = 'actual' THEN NOW() ELSE last_actual_sample_at END
                WHERE pattern_key = %s AND department = %s
            """, (new_layout, new_detail, new_doc, new_total,
                  m2_layout, m2_detail, m2_doc, m2_total,
                  new_occ, confidence, source, source, pattern_key, dept))
            
            logger.debug(f"   ✅ UPDATED: '{name[:40]}' occ: {old_occ}→{new_occ}, total: {old_total:.2f}→{new_total:.2f}h")
            
        else:
            # INSERT new pattern
            confidence = 0.1
            
            cur.execute("""
                INSERT INTO component_patterns (
                    name, pattern_key, department,
                    avg_hours_3d_layout, avg_hours_3d_detail, avg_hours_2d, avg_hours_total,
                    m2_layout, m2_detail, m2_doc, m2_total,
                    occurrences, confidence, source, 
                    last_updated,
                    last_actual_sample_at
                ) VALUES (
                    %s, %s, %s,
                    %s, %s, %s, %s,
                    0, 0, 0, 0,
                    1, %s, %s,
                    NOW(),
                    CASE WHEN %s = 'actual' THEN NOW() ELSE NULL END
                )
            """, (name, pattern_key, dept, 
                  layout_h, detail_h, doc_h, total_h,
                  confidence, source, source))
            
            logger.debug(f"   ✅ INSERTED: '{name[:40]}' total: {total_h:.2f}h")
        
        return True

    except Exception as e:
        logger.error(f"❌ update_pattern_smart ERROR for '{name[:40]}': {e}", exc_info=True)
        return False


def update_category_baseline(cur, dept, category, layout_h, detail_h, doc_h):
    """Aktualizuje baseline kategorii (średnie ruchome metodą Welforda)."""
    cur.execute("""
    SELECT mean_layout, mean_detail, mean_doc, m2_layout, m2_detail, m2_doc, occurrences
    FROM category_baselines WHERE department=%s AND category=%s
    """, (dept, category))
    row = cur.fetchone()

    if row:
        ml, md, mc, m2l, m2d, m2c, n0 = row
        ml, m2l, _ = _welford_step(ml, m2l, n0, float(layout_h))
        md, m2d, _ = _welford_step(md, m2d, n0, float(detail_h))
        mc, m2c, _ = _welford_step(mc, m2c, n0, float(doc_h))
        n1 = (n0 or 0) + 1
        conf = min(1.0, n1 / 10.0)
        cur.execute("""
            UPDATE category_baselines
            SET mean_layout=%s, mean_detail=%s, mean_doc=%s,
                m2_layout=%s, m2_detail=%s, m2_doc=%s,
                occurrences=%s, confidence=%s, last_updated=NOW()
            WHERE department=%s AND category=%s
        """, (ml, md, mc, m2l, m2d, m2c, n1, conf, dept, category))
    else:
        cur.execute("""
            INSERT INTO category_baselines (department, category, mean_layout, mean_detail, mean_doc, occurrences, confidence)
            VALUES (%s, %s, %s, %s, %s, 1, 0.1)
        """, (dept, category, layout_h, detail_h, doc_h))

def update_bundle(cur, dept: str, parent_name: str, sub_name: str, qty: int):
    """Relacja parent→sub w component_bundles + sumy ilości i occurrences."""
    try:
        pkey = canonicalize_name(parent_name or "")
        skey = canonicalize_name(sub_name or "")
        if not pkey or not skey:
            return
        q = max(1, int(qty or 1))
        cur.execute("""
            INSERT INTO component_bundles (department, parent_key, parent_name, sub_key, sub_name, occurrences, total_qty, confidence)
            VALUES (%s, %s, %s, %s, %s, 1, %s, 0.1)
            ON CONFLICT (department, parent_key, sub_key)
            DO UPDATE SET
                occurrences = component_bundles.occurrences + 1,
                total_qty = component_bundles.total_qty + EXCLUDED.total_qty,
                confidence = LEAST(1.0, (component_bundles.occurrences + 1) / 10.0),
                last_updated = NOW()
        """, (dept, pkey, parent_name, skey, sub_name, float(q)))
    except Exception as e:
        logger.warning(f"update_bundle err: {e}")

def learn_from_historical_components(cur, dept: str, components: list, distribute: str = 'qty'):
    """
    Uczy:
    - component_patterns: komponenty główne + sub-komponenty (wg qty lub po równo),
    - component_bundles: częste pary parent→sub.
    """
    # ════════════════════════════════════════════════════════════
    # DODAJ TO NA POCZĄTKU FUNKCJI:
    # ════════════════════════════════════════════════════════════
    learned_patterns = 0
    learned_bundles = 0
    skipped_summary = 0
    skipped_no_hours = 0
    skipped_no_name = 0
    # ════════════════════════════════════════════════════════════
    for comp in components or []:
        try:
            name = comp.get('name', '')

            # Pomiń placeholder names i puste nazwy
            if not name or name in ['[part]', '[assembly]', ' ']:
                skipped_no_name += 1
                continue
            is_summary = bool(comp.get('is_summary'))
            subs = comp.get('subcomponents', []) or []

            # Bundles – zawsze
            for sub in subs:
                update_bundle(cur, dept, name, sub.get('name', ''), sub.get('quantity', 1))
                learned_bundles += 1  # <-- DODAJ

            # Pomijamy “sumaryczne”
            if is_summary:
                skipped_summary += 1  # <-- DODAJ
                continue

            layout = float(comp.get('hours_3d_layout', 0) or 0)
            detail = float(comp.get('hours_3d_detail', 0) or 0)
            doc = float(comp.get('hours_2d', 0) or 0)
            total = layout + detail + doc

            # Wzorzec główny
            if total > 0:
                update_pattern_smart(cur, name, dept, layout, detail, doc, source='historical_excel')
                learned_patterns += 1  # <-- DODAJ
            else:
                skipped_no_hours += 1  # <-- DODAJ

            # Rozdział na suby
            if subs and total > 0:
                if distribute == 'qty':
                    total_qty = sum(int(s.get('quantity', 1) or 1) for s in subs) or len(subs)
                    for sub in subs:
                        q = max(1, int(sub.get('quantity', 1) or 1))
                        w = (q / total_qty) if total_qty else (1.0 / len(subs))
                        sl, sd, sdoc = layout * w, detail * w, doc * w
                        update_pattern_smart(cur, sub.get('name', ''), dept, sl, sd, sdoc, source='historical_excel_sub')
                        learned_patterns += 1  # <-- DODAJ
                else:
                    n = len(subs)
                    if n > 0:
                        w = 1.0 / n
                        for sub in subs:
                            sl, sd, sdoc = layout * w, detail * w, doc * w
                            update_pattern_smart(cur, sub.get('name', ''), dept, sl, sd, sdoc, source='historical_excel_sub')
                            learned_patterns += 1  # <-- DODAJ
        except Exception as e:
            logger.warning(f"learn_from_historical_components err for '{comp.get('name','?')}': {e}")

 # ════════════════════════════════════════════════════════════
    # DODAJ TO NA KOŃCU FUNKCJI:
    # ════════════════════════════════════════════════════════════
    logger.info(f"""
    📊 STATYSTYKI UCZENIA dla działu {dept}:
       ✅ Wzorców nauczonych: {learned_patterns}
       ✅ Bundles nauczonych: {learned_bundles}
       ⏭️  Pominięto (summary): {skipped_summary}
       ⏭️  Pominięto (brak godzin): {skipped_no_hours}
       ⏭️  Pominięto (brak nazwy): {skipped_no_name}
    """)
    # ════════════════════════════════════════════════════════════

# === HEURYSTYKI ===
HEURISTIC_LIBRARY = [
    (['docisk', 'clamp'], 0.5, 1.5, 0.5),
    (['śruba trapezowa', 'trapez'], 0.2, 0.8, 0.3),
    (['konsola', 'bracket'], 0.3, 1.0, 0.4),
    (['płyta', 'plate'], 0.2, 0.7, 0.4),
]

def heuristic_estimate_for_name(name: str):
    n = name.lower()
    for keys, l, d, doc in HEURISTIC_LIBRARY:
        if any(k in n for k in keys):
            return l, d, doc, f"Heurystyka: {', '.join(keys)}"
    return 0.0, 0.0, 0.0, ""

# === PROPOZYCJE DODATKÓW (patterns + heurystyki) ===
def propose_adjustments_for_components(conn, components, department, conservativeness=1.0, sim_threshold=0.6):
    proposals = []
    for comp in components:
        subs = comp.get('subcomponents', [])
        if not subs:
            continue

        # agregacja qty
        agg = {}
        for s in subs:
            qty = int(s.get('quantity', 1) or 1)
            nm = s.get('name', '').strip()
            if not nm:
                continue
            key = canonicalize_name(nm)
            if key not in agg:
                agg[key] = {'display_name': nm, 'qty': 0}
            agg[key]['qty'] += qty

        adds = []
        for key_name, info in agg.items():
            display_name = info['display_name']
            qty = info['qty']

            # 1) wzorzec
            similar = find_similar_components(conn, display_name, department, limit=1)
            used = False
            if similar:
                s0 = similar[0]
                sim = float(s0.get('similarity') or 0.0)
                if sim >= sim_threshold:
                    l = float(s0.get('avg_hours_3d_layout') or 0.0)
                    d = float(s0.get('avg_hours_3d_detail') or 0.0)
                    dc = float(s0.get('avg_hours_2d') or 0.0)
                    tot = float(s0.get('avg_hours_total') or (l + d + dc))
                    if tot > 0 and (l + d + dc) == 0:
                        l, d, dc = tot * 0.3, tot * 0.5, tot * 0.2
                    adds.append({
                        "name": display_name, "qty": qty,
                        "layout_add": l * qty * conservativeness,
                        "detail_add": d * qty * conservativeness,
                        "doc_add": dc * qty * conservativeness,
                        "reason": f"Wzorzec: {s0.get('name')} (sim={sim*100:.0f}%)",
                        "source": "pattern",
                        "confidence": float(s0.get('confidence') or 0.5)
                    })
                    used = True

            # 2) heurystyka
            if not used:
                l, d, dc, why = heuristic_estimate_for_name(display_name)
                if l + d + dc > 0:
                    adds.append({
                        "name": display_name, "qty": qty,
                        "layout_add": l * qty * conservativeness,
                        "detail_add": d * qty * conservativeness,
                        "doc_add": dc * qty * conservativeness,
                        "reason": why or "Heurystyka ogólna",
                        "source": "heuristic",
                        "confidence": 0.4
                    })
        if adds:
            proposals.append({"parent": comp.get('name', 'bez nazwy'), "adds": adds})
    return proposals

# === PROPOZYCJE Z HISTORII (bundles) ===
def propose_bundles_for_component(conn, parent_name: str, department: str,
                                  conservativeness: float = 1.0,
                                  top_k: int = 5, min_occ: int = 2) -> list:
    pkey = canonicalize_name(parent_name or "")
    if not pkey:
        return []

    rows = []
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT sub_name, occurrences, total_qty, confidence
                FROM component_bundles
                WHERE department=%s AND parent_key=%s AND occurrences >= %s
                ORDER BY occurrences DESC
                LIMIT %s
            """, (department, pkey, min_occ, top_k))
            rows = cur.fetchall() or []
    except Exception as e:
        logger.warning(f"propose_bundles_for_component query err: {e}")
        return []

    proposals = []
    for r in rows:
        occ = max(1, int(r.get('occurrences') or 1))
        total_qty = float(r.get('total_qty') or 0.0)
        typical_qty = int(round(total_qty / occ)) if total_qty > 0 else 1
        typical_qty = max(1, typical_qty)

        # dociągnięcie wzorca godzin
        try:
            similar = find_similar_components(conn, r['sub_name'], department, limit=1)
        except Exception:
            similar = []

        if similar:
            s0 = similar[0]
            l = float(s0.get('avg_hours_3d_layout') or 0.0)
            d = float(s0.get('avg_hours_3d_detail') or 0.0)
            dc = float(s0.get('avg_hours_2d') or 0.0)
            tot = float(s0.get('avg_hours_total') or (l + d + dc))
            if tot > 0 and (l + d + dc) == 0:
                l, d, dc = tot * 0.3, tot * 0.5, tot * 0.2

            proposals.append({
                "name": r['sub_name'],
                "qty": typical_qty,
                "layout_add": l * typical_qty * conservativeness,
                "detail_add": d * typical_qty * conservativeness,
                "doc_add": dc * typical_qty * conservativeness,
                "reason": f"Historia: często występuje z {parent_name} (occ={occ})",
                "source": "bundle",
                "confidence": float(r.get('confidence') or 0.5)
            })
        else:
            # fallback do heurystyki
            l, d, dc, why = heuristic_estimate_for_name(r['sub_name'])
            if l + d + dc > 0:
                proposals.append({
                    "name": r['sub_name'],
                    "qty": typical_qty,
                    "layout_add": l * typical_qty * conservativeness,
                    "detail_add": d * typical_qty * conservativeness,
                    "doc_add": dc * typical_qty * conservativeness,
                    "reason": why or f"Historia (bundle) bez wzorca",
                    "source": "bundle_heuristic",
                    "confidence": 0.35
                })

    return proposals
# === Batch import historycznych Exceli (z opcją uczenia) ===                                      
def batch_import_excels(files, department: str,
                        learn_from_import: bool = False,
                        distribute: str = 'qty'):
    """
    Batch import historycznych plików Excel:
    - parsuje komponenty + komentarze (openpyxl comments),
    - opis pobierany AUTOMATYCZNIE z A1 pierwszej zakładki,
    - zapisuje projekt jako Excel-only (is_historical/locked),
    - od razu generuje embedding opisu (pgvector),
    - opcjonalnie uczy wzorce (patterns) i bundles.
    Zwraca listę wyników: {file, status, project_id?, hours?, desc?, error?}
    """
    results = []
    with get_db_connection() as conn, conn.cursor() as cur:
        for f in files:
            try:
                fname = getattr(f, "name", "import.xlsx")
                proj_name = os.path.splitext(os.path.basename(fname))[0]

                # Wczytaj bytes raz i wykorzystaj dwa razy (parser + A1)
                content = f.read()
                bio_parse = BytesIO(content)
                bio_a1 = BytesIO(content)

                # 1) Parser (wartości + komentarze/note)
                parsed = parse_cad_project_structured_with_xlsx_comments(bio_parse)
                comps_full = parsed.get('components', []) or []
                totals = parsed.get('totals', {}) or {}

                est_l = float(totals.get('layout', 0) or 0)
                est_d = float(totals.get('detail', 0) or 0)
                est_doc = float(totals.get('documentation', 0) or 0)
                est_total = float(totals.get('total', est_l + est_d + est_doc) or 0)

                # 2) Opis z A1 (pierwsza zakładka)
                description = extract_scope_from_excel_a1_first_sheet(bio_a1).strip()
                if not description:
                    description = f"Projekt historyczny: {proj_name}."

                # 3) Zapisz projekt jako historyczny, zablokowany dla AI
                cur.execute("""
                    INSERT INTO projects (
                        name, client, department, description, components,
                        estimated_hours_3d_layout, estimated_hours_3d_detail, estimated_hours_2d,
                        estimated_hours, ai_analysis,
                        is_historical, estimation_mode, totals_source, locked_totals
                    ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                             TRUE, 'excel_only', 'excel', TRUE)
                    RETURNING id
                """, (
                    proj_name, None, department, description,
                    json.dumps(comps_full, ensure_ascii=False),
                    est_l, est_d, est_doc, est_total,
                    '[HISTORICAL_IMPORT]'
                ))
                pid = cur.fetchone()[0]

                # 4) Embedding opisu (pgvector)
                ensure_project_embedding(cur, pid, description)

                # 5) Uczenie wzorców/bundles (opcjonalnie)
                if learn_from_import and comps_full:
                    logger.info(f"🧠 UCZĘ WZORCE: {len(comps_full)} komponentów z działu {department}")  # <-- DODAJ TO
                    learn_from_historical_components(cur, department, comps_full, distribute=distribute)
                    logger.info(f"✅ UCZENIE ZAKOŃCZONE")  # <-- DODAJ TO

                # ════════════════════════════════════════════════════════════
                # DODAJ TO BEZPOŚREDNIO PO learn_from_historical_components:
                # ════════════════════════════════════════════════════════════
                # Sprawdź ile wzorców w bazie PRZED commit
                cur.execute("SELECT COUNT(*) FROM component_patterns WHERE department=%s", (department,))
                pattern_count = cur.fetchone()[0]
                logger.info(f"📊 Po uczeniu, przed commit: {pattern_count} wzorców w działu {department}")
                # ════════════════════════════════════════════════════════════
                
                conn.commit()
                
                # ════════════════════════════════════════════════════════════
                # DODAJ TO BEZPOŚREDNIO PO conn.commit():
                # ════════════════════════════════════════════════════════════
                # Sprawdź ile wzorców w bazie PO commit
                cur.execute("SELECT COUNT(*) FROM component_patterns WHERE department=%s", (department,))
                pattern_count_after = cur.fetchone()[0]
                logger.info(f"📊 Po commit: {pattern_count_after} wzorców w działu {department}")
                # ════════════════════════════════════════════════════════════
                    
                
                results.append({
                    "file": fname,
                    "status": "success",
                    "project_id": pid,
                    "hours": est_total,
                    "desc": (description[:120] + ("..." if len(description) > 120 else ""))
                })
            except Exception as e:
                conn.rollback()
                logger.exception("Batch import error")
                results.append({"file": getattr(f, 'name', 'unknown'), "status": "error", "error": str(e)})
    return results
                            
def enhance_estimation_with_web(component_name: str, department: str, enable_web: bool) -> dict:
    """
    Wzbogaca estymację komponentu o dane z sieci (normy, benchmarki).
    Zwraca: {"norms": [...], "typical_hours": float, "notes": "..."}
    """
    if not enable_web or not st.session_state.get("allow_web_lookup"):
        return {}
    
    # Przykład: wyszukaj normy dla komponentu
    try:
        from duckduckgo_search import DDGS
        import trafilatura
    except ImportError:
        return {"error": "duckduckgo-search/trafilatura not installed"}
    
    results = {}
    
    # Zapytanie: normy dla typu komponentu
    query = f"{component_name} CAD standard ISO EN time estimation"
    try:
        with DDGS() as ddg:
            hits = ddg.text(query, region="en-us", safesearch="moderate", max_results=2)
            for h in hits:
                url = h.get("href")
                if url:
                    try:
                        content = trafilatura.fetch_url(url)
                        text = trafilatura.extract(content) or ""
                        if text:
                            results["web_context"] = text[:500]  # Pierwsze 500 znaków
                            break
                    except Exception:
                        continue
    except Exception as e:
        logger.warning(f"Web search failed: {e}")
    
    return results
# === FUNKCJE POMOCNICZE DLA AI BRIEF I ANALIZY ===

def build_brief_prompt(description: str, components: list, pdf_text: str, department: str) -> str:
    """
    Buduje prompt do generowania briefu projektu.
    
    Args:
        description: Opis projektu od użytkownika
        components: Lista komponentów z Excela/JSON
        pdf_text: Tekst z plików PDF
        department: Kod działu (131-135)
    
    Returns:
        Sformatowany prompt dla AI
    """
    # Przykładowe komponenty (max 10)
    comp_names = [c.get('name', '') for c in components[:10] if not c.get('is_summary', False)]
    comp_list = "\n".join([f"- {name}" for name in comp_names if name]) or "Brak komponentów"
    
    # Kontekst branżowy
    context = DEPARTMENT_CONTEXT.get(department, "")
    
    return f"""Jesteś ekspertem CAD i project managerem. Przeanalizuj projekt i stwórz szczegółowy brief.

DZIAŁ: {department}
{context}

OPIS PROJEKTU:
{description[:1500] if description else "Brak opisu"}

PRZYKŁADOWE KOMPONENTY:
{comp_list}

SPECYFIKACJE TECHNICZNE:
{pdf_text[:2500] if pdf_text else "Brak dodatkowych specyfikacji"}

ZADANIE: Zwróć szczegółowy brief projektu w formacie JSON.

WYMAGANA STRUKTURA JSON:
{{
  "brief_md": "Krótki opis projektu (2-3 akapity w Markdown) - co to za projekt, główne wymagania, złożoność",
  "scope": ["zakres prac 1", "zakres prac 2", "zakres prac 3"],
  "assumptions": ["założenie techniczne 1", "założenie 2"],
  "missing_info": ["brakująca informacja 1", "pytanie do klienta 2"],
  "risks": [
    {{"risk": "opis ryzyka", "impact": "wysoki/średni/niski", "mitigation": "jak zminimalizować"}},
    {{"risk": "inne ryzyko", "impact": "średni", "mitigation": "plan mitygacji"}}
  ],
  "checklist": ["punkt kontrolny 1", "punkt kontrolny 2", "weryfikacja 3"],
  "open_questions": ["pytanie do zespołu 1", "pytanie techniczne 2"]
}}

ZASADY:
- Pisz TYLKO po polsku
- Zwróć WYŁĄCZNIE JSON (bez komentarzy, bez tekstu przed/po)
- W "risks" KAŻDE ryzyko MUSI mieć: risk, impact, mitigation
- brief_md może zawierać Markdown (nagłówki ##, listy, pogrubienia **)
- Bądź konkretny i techniczny
"""


def parse_brief_response(resp_text: str) -> dict:
    """
    Parsuje odpowiedź AI z briefem projektu.
    
    Args:
        resp_text: Surowa odpowiedź od AI (może zawierać code fences)
    
    Returns:
        Słownik z brieFem lub struktura zastępcza
    """
    try:
        # Usuń code fences (```json ... ```)
        clean = resp_text.strip()
        if clean.startswith("```json"):
            clean = clean[7:]
        if clean.startswith("```"):
            clean = clean[3:]
        if clean.endswith("```"):
            clean = clean[:-3]
        clean = clean.strip()
        
        # Parsuj JSON
        data = json.loads(clean)
        
        # Walidacja struktury
        required_keys = ["brief_md", "scope", "assumptions", "missing_info", "risks", "checklist", "open_questions"]
        for key in required_keys:
            if key not in data:
                data[key] = [] if key != "brief_md" else ""
        
        # Walidacja ryzyk (MUSZĄ mieć risk, impact, mitigation)
        validated_risks = []
        for r in data.get("risks", []):
            if isinstance(r, dict) and all(k in r for k in ["risk", "impact", "mitigation"]):
                validated_risks.append(r)
        data["risks"] = validated_risks
        
        return data
        
    except json.JSONDecodeError as e:
        logger.error(f"Brief JSON parsing error: {e}")
        # Fallback - zwróć surowy tekst jako brief
        return {
            "brief_md": f"**Błąd parsowania JSON**\n\n{resp_text[:800]}",
            "scope": [],
            "assumptions": [],
            "missing_info": ["Nie udało się sparsować odpowiedzi AI"],
            "risks": [],
            "checklist": [],
            "open_questions": []
        }
    except Exception as e:
        logger.error(f"Brief parsing error: {e}")
        return {
            "brief_md": f"**Błąd:** {str(e)}",
            "scope": [],
            "assumptions": [],
            "missing_info": [],
            "risks": [],
            "checklist": [],
            "open_questions": []
        }
def extract_keywords(text: str) -> list:
    """
    Wyciąga kluczowe słowa techniczne z opisu.
    Filtruje stopwords i zostawia tylko rzeczowniki techniczne.
    """
    stopwords = {'i', 'a', 'z', 'do', 'w', 'na', 'dla', 'o', 'po', 'ze', 'od', 
                 'the', 'and', 'or', 'of', 'to', 'in', 'on', 'at', 'from', 
                 'jest', 'są', 'będzie', 'ma', 'będą', 'można', 'tego', 'tej',
                 'tym', 'ten', 'ta', 'to', 'jak', 'się', 'już', 'tylko'}
    
    # Tokenizacja
    words = re.findall(r'\b\w+\b', text.lower())
    
    # Filtruj
    keywords = [w for w in words if len(w) > 3 and w not in stopwords]
    
    # Usuń duplikaty zachowując kolejność
    seen = set()
    unique = []
    for k in keywords:
        if k not in seen:
            seen.add(k)
            unique.append(k)
    
    return unique[:10]  # Max 10 słów kluczowych
def intelligent_decomposition(description: str, department: str, conn) -> dict:
    """
    Inteligentna dekompozycja opisu na komponenty używając:
    1. Wzorców z bazy (semantic search)
    2. Podobnych projektów (semantic search)
    3. Typowych zestawów komponentów (bundles)
    
    Zwraca: {
        "suggested_components": [...],
        "context_from_db": "...",
        "similar_projects": [...]
    }
    """
    result = {
        "suggested_components": [],
        "context_from_db": "",
        "similar_projects": []
    }
    
    # 1. Wyciągnij kluczowe terminy z opisu
    keywords = extract_keywords(description)
    logger.info(f"🔍 Extracted keywords: {keywords}")
    
    # ═══════════════════════════════════════════════════════════
    # 2. Dla każdego słowa kluczowego znajdź wzorce
    # ═══════════════════════════════════════════════════════════
    all_patterns = []
    for keyword in keywords:
        # Semantic search po wzorcach (find_similar_components już ma własny cursor)
        similar = find_similar_components(conn, keyword, department, limit=3)
        all_patterns.extend(similar)
    
    # ═══════════════════════════════════════════════════════════
    # 3. Deduplikuj wzorce
    # ═══════════════════════════════════════════════════════════
    seen = set()
    unique_patterns = []
    for p in all_patterns:
        key = canonicalize_name(p.get('name', ''))
        if key not in seen:
            seen.add(key)
            unique_patterns.append(p)
    
    logger.info(f"🧠 Found {len(unique_patterns)} unique patterns from DB")
    
    # ═══════════════════════════════════════════════════════════
    # 4. Znajdź podobne projekty
    # ═══════════════════════════════════════════════════════════
    similar_projects = find_similar_projects_semantic(conn, description, department, limit=3)
    result["similar_projects"] = similar_projects
    
    # ═══════════════════════════════════════════════════════════
    # 5. Zbuduj kontekst dla AI
    # ═══════════════════════════════════════════════════════════
    context = "═══════════════════════════════════════════════════════════\n"
    context += "KOMPONENTY ZNALEZIONE W BAZIE (użyj ich w dekompozycji!):\n"
    context += "═══════════════════════════════════════════════════════════\n\n"
    
    for p in unique_patterns[:15]:
        context += f"- **{p['name']}**: "
        context += f"Layout {p.get('avg_hours_3d_layout', 0):.1f}h, "
        context += f"Detail {p.get('avg_hours_3d_detail', 0):.1f}h, "
        context += f"Doc {p.get('avg_hours_2d', 0):.1f}h "
        context += f"(confidence: {p.get('confidence', 0):.2f}, n={p.get('occurrences', 0)})\n"
        
        # ═══════════════════════════════════════════════════════════
        # 6. Znajdź typowe zestawy (bundles) dla każdego wzorca
        # ═══════════════════════════════════════════════════════════
        try:
            bundle_adds = propose_bundles_for_component(
                conn, p['name'], department, 
                conservativeness=1.0, top_k=3, min_occ=2
            )
            if bundle_adds:
                bundle_names = [a['name'] for a in bundle_adds[:3]]
                context += f"  └─ Typowo występuje z: {', '.join(bundle_names)}\n"
        except Exception as e:
            logger.warning(f"Bundle lookup failed for '{p['name']}': {e}")
    
    # ═══════════════════════════════════════════════════════════
    # 7. Dodaj podobne projekty do kontekstu
    # ═══════════════════════════════════════════════════════════
    if similar_projects:
        context += "\n═══════════════════════════════════════════════════════════\n"
        context += "PODOBNE PROJEKTY W BAZIE:\n"
        context += "═══════════════════════════════════════════════════════════\n\n"
        
        for proj in similar_projects:
            context += f"- **{proj['name']}** ({proj['client'] or 'N/A'}): "
            context += f"{proj['estimated_hours']:.1f}h "
            context += f"(similarity: {proj.get('similarity', 0)*100:.0f}%)\n"
    
    result["context_from_db"] = context
    result["suggested_components"] = unique_patterns
    
    return result

def extract_keywords(text: str) -> list:
    """
    Wyciąga kluczowe słowa techniczne z opisu.
    Filtruje stopwords i zostawia tylko rzeczowniki techniczne.
    """
    # Podstawowa lista słów kluczowych (możesz rozbudować)
    stopwords = {'i', 'a', 'z', 'do', 'w', 'na', 'dla', 'o', 'po', 'ze', 'od', 
                 'the', 'and', 'or', 'of', 'to', 'in', 'on', 'at', 'from'}
    
    # Tokenizacja
    words = re.findall(r'\b\w+\b', text.lower())
    
    # Filtruj
    keywords = [w for w in words if len(w) > 3 and w not in stopwords]
    
    # Usuń duplikaty zachowując kolejność
    seen = set()
    unique = []
    for k in keywords:
        if k not in seen:
            seen.add(k)
            unique.append(k)
    
    return unique[:10]  # Max 10 słów kluczowych

def build_analysis_prompt(description: str, components: list, 
                          learned_patterns: list, pdf_text: str, 
                          department: str, conn=None) -> str:
    """
    Buduje prompt do analizy komponentów i estymacji godzin.
    Jeśli conn podane - użyje inteligentnej dekompozycji z bazy.
    """
    
    # ═══════════════════════════════════════════════════════════
    # DEBUG
    # ═══════════════════════════════════════════════════════════
    logger.info(f"🔍 build_analysis_prompt INPUTS:")
    logger.info(f"   📝 description: '{description[:100]}...' (len={len(description)})")
    logger.info(f"   📦 components: {len(components)} items")
    logger.info(f"   🧠 learned_patterns: {len(learned_patterns)} items")
    logger.info(f"   📄 pdf_text: {len(pdf_text)} chars")
    logger.info(f"   🏢 department: {department}")
    
    # ═══════════════════════════════════════════════════════════
    # INTELIGENTNA DEKOMPOZYCJA Z BAZY
    # ═══════════════════════════════════════════════════════════
    db_context = ""
    if conn and description:
        logger.info("🧠 Uruchamiam intelligent_decomposition...")
        decomp = intelligent_decomposition(description, department, conn)
        db_context = decomp["context_from_db"]
        logger.info(f"✅ Znaleziono {len(decomp['suggested_components'])} sugerowanych komponentów")
    
    # Kontekst branżowy
    context = DEPARTMENT_CONTEXT.get(department, "")
    
    # Przykładowe komponenty z Excela/JSON (jeśli są)
    comp_examples = []
    for c in components[:30]:
        if not c.get('is_summary', False):
            name = c.get('name', 'Bez nazwy')
            layout = c.get('hours_3d_layout', 0)
            detail = c.get('hours_3d_detail', 0)
            doc = c.get('hours_2d', 0)
            comment = c.get('comment', '')
            subs = c.get('subcomponents', [])
            
            line = f"- **{name}**: Layout {layout:.1f}h, Detail {detail:.1f}h, 2D {doc:.1f}h"
            
            if comment:
                line += f"\n  └─ Uwagi: {comment[:100]}"
            
            if subs:
                line += f"\n  └─ Zawiera: "
                sub_names = [f"{s.get('quantity',1)}x {s.get('name','')}" for s in subs[:5]]
                line += ", ".join(sub_names)
                if len(subs) > 5:
                    line += f" ... (+{len(subs)-5})"
            
            comp_examples.append(line)
    
    comp_str = "\n".join(comp_examples) if comp_examples else "(Brak przykładów z Excela)"
    
    # ═══════════════════════════════════════════════════════════
    # PROMPT Z KONTEKSTEM Z BAZY
    # ═══════════════════════════════════════════════════════════
    
    return f"""{MASTER_PROMPT}

KONTEKST PROJEKTU:

DZIAŁ: {department}
{context}

{db_context}

OPIS UŻYTKOWNIKA:
{description[:2000] if description else "Brak szczegółowego opisu"}

KOMPONENTY Z EXCELA/JSON (referencyjne - opcjonalne):
{comp_str}

SPECYFIKACJE/PDF (opcjonalne):
{pdf_text[:2500] if pdf_text else "Brak dodatkowych specyfikacji"}

═══════════════════════════════════════════════════════════
ZADANIE - INTELIGENTNA DEKOMPOZYCJA:
═══════════════════════════════════════════════════════════

1. Przeanalizuj OPIS UŻYTKOWNIKA
2. Sprawdź KOMPONENTY ZNALEZIONE W BAZIE powyżej
3. Jeśli opis wspomina o elementach podobnych do tych z bazy (np. "wspornik", "czujnik", "płyta"):
   → MUSISZ dodać je jako osobne pozycje w "components"
4. Jeśli opis wspomina o ilościach (np. "4x wspornik"):
   → Uwzględnij to w dekompozycji (4 osobne komponenty lub 1 z qty)
5. Użyj godzin z bazy jako punktu odniesienia, ale dostosuj do specyfiki projektu
6. Jeśli nie masz pewności co do komponentu - lepiej dodać niż pominąć

PRZYKŁAD DOBREJ DEKOMPOZYCJI:

OPIS: "Stacja dociskania z 4 wspornikami, 2 czujnikami i płytą bazową"

BAZA ZAWIERA:
- wspornik: 2.5h
- czujnik: 1.8h  
- płyta: 5.0h

PRAWIDŁOWA ODPOWIEDŹ:
{{
  "components": [
    {{"name": "Rama główna stacji", "layout_h": 3.0, "detail_h": 8.0, "doc_h": 4.0}},
    {{"name": "Wspornik montażowy", "layout_h": 0.6, "detail_h": 1.5, "doc_h": 0.4}},
    {{"name": "Wspornik montażowy", "layout_h": 0.6, "detail_h": 1.5, "doc_h": 0.4}},
    {{"name": "Wspornik montażowy", "layout_h": 0.6, "detail_h": 1.5, "doc_h": 0.4}},
    {{"name": "Wspornik montażowy", "layout_h": 0.6, "detail_h": 1.5, "doc_h": 0.4}},
    {{"name": "Czujnik pozycji", "layout_h": 0.4, "detail_h": 1.0, "doc_h": 0.4}},
    {{"name": "Czujnik pozycji", "layout_h": 0.4, "detail_h": 1.0, "doc_h": 0.4}},
    {{"name": "Płyta bazowa", "layout_h": 1.2, "detail_h": 3.0, "doc_h": 0.8}}
  ],
  ...
}}

Przeanalizuj dokładnie i zwróć JSON z PEŁNĄ dekompozycją.
"""

# ═══════════════════════════════════════════════════════════
# SPRINT 1: PYTANIA DOPRECYZOWUJĄCE
# ═══════════════════════════════════════════════════════════

def generate_clarifying_questions(description: str, department: str, pdf_text: str = "") -> list:
    """
    Generuje pytania doprecyzowujące na podstawie opisu projektu.
    Używa AI do identyfikacji brakujących informacji.

    Zwraca: [{"question": "...", "why": "..."}, ...]
    """
    if not description or len(description.strip()) < 50:
        return []

    prompt = f"""Jesteś senior konstruktorem CAD. Przeanalizuj poniższy opis projektu i wygeneruj 3-5 KLUCZOWYCH pytań, które pomogą w dokładniejszej estymacji czasu.

DZIAŁ: {DEPARTMENTS.get(department, department)}

OPIS PROJEKTU:
{description[:1000]}

PDF/SPECYFIKACJA:
{pdf_text[:500] if pdf_text else "Brak"}

ZASADY:
- Pytaj TYLKO o rzeczy krytyczne dla estymacji (materiały, ilości, normy, procesy)
- NIE pytaj o rzeczy oczywiste lub już opisane
- Maksymalnie 5 pytań
- Każde pytanie musi mieć uzasadnienie (dlaczego to ważne)

Zwróć JSON:
{{
  "questions": [
    {{
      "question": "Krótkie pytanie?",
      "why": "Dlaczego to ważne dla estymacji"
    }}
  ]
}}

PRZYKŁAD:
Opis: "Rama spawana z wspornikami"

ODPOWIEDŹ:
{{
  "questions": [
    {{
      "question": "Jaki materiał ramy? (S235JR, S355J2, Aluminium)",
      "why": "Materiał wpływa na czas obróbki (+20% dla S355) i spawanie"
    }},
    {{
      "question": "Ile wsporników?",
      "why": "Bezpośrednio wpływa na czas realizacji (każdy ~2-3h)"
    }},
    {{
      "question": "Jakie wymiary ramy? (dł x szer x wys w mm)",
      "why": "Duże wymiary (>5m) zwiększają złożoność o 25%"
    }}
  ]
}}

Zwróć TYLKO JSON, bez tekstu.
"""

    try:
        ai_model = st.session_state.get("selected_text_model", "qwen2.5:7b")
        logger.info(f"🤔 Generuję pytania doprecyzowujące przez {ai_model}...")

        response = query_ollama(prompt, model=ai_model, format_json=True)

        # Parse JSON
        data = safe_json_loads(response)
        questions = data.get("questions", [])

        logger.info(f"✅ Wygenerowano {len(questions)} pytań")
        return questions[:5]  # Max 5 pytań

    except Exception as e:
        logger.error(f"❌ Nie udało się wygenerować pytań: {e}", exc_info=True)
        return []

def render_new_project_page():
    st.header("🆕 Nowy Projekt")


    department = st.selectbox(
        "Wybierz dział*",
        options=list(DEPARTMENTS.keys()),
        format_func=lambda x: f"{x} - {DEPARTMENTS[x]}",
        key="department"
    )

    st.info(f"📋 {DEPARTMENT_CONTEXT[department]}")

    col1, col2 = st.columns(2)
    with col1:
        st.text_input("Nazwa projektu*", key="project_name")
        st.text_input("Klient", key="client")

        # SPRINT 1: Monitoruj zmiany w opisie i resetuj pytania
        current_desc = st.text_area("Opis", height=200, key="description")
        if current_desc != st.session_state.get("_prev_description"):
            st.session_state["_prev_description"] = current_desc
            st.session_state["questions_answered"] = False
            st.session_state["clarifying_answers"] = {}

    with col2:
        excel_file = st.file_uploader("Excel", type=['xlsx', 'xls'])
        image_files = st.file_uploader("Zdjęcia/Rysunki", type=['jpg', 'png'], accept_multiple_files=True)
        pdf_files = st.file_uploader("PDF", type=['pdf'], accept_multiple_files=True)
        json_files = st.file_uploader("JSON (doc-converter/AI)", type=['json'], accept_multiple_files=True)
    pasted_text = st.text_area("Dodatkowy tekst/specyfikacja (wklej – opcjonalnie)", height=120, key="pasted_text")

    # 🔹 AI Brief: opis zadania i checklista
    st.subheader("📝 AI: Opis zadania i checklista")
    if st.button("📝 Generuj opis zadania (AI)", use_container_width=True):
        # Komponenty z Excela (przykłady)
        components_for_brief = []
        if excel_file is not None:
            try:
                components_for_brief = parse_cad_project_structured_with_xlsx_comments(BytesIO(excel_file.getvalue()))['components']
            except Exception:
                components_for_brief = []

        # Komponenty z JSON (doc-converter/AI)
        components_from_json_for_brief = []
        if json_files:
            for jf in json_files:
                try:
                    data = safe_json_loads(jf.getvalue())
                    components_from_json_for_brief += parse_components_from_docconv_json(data)
                except Exception:
                    pass

        pdf_text_for_brief = ""
        if pdf_files:
            pdf_text_for_brief = "\n".join([extract_text_from_pdf(pf) for pf in pdf_files])

        if st.session_state.get("pasted_text"):
            pdf_text_for_brief = (pdf_text_for_brief + "\n\n" + st.session_state.get("pasted_text")).strip()

        # Sprawdź czy są dane wejściowe
        if not st.session_state.get("description") and not components_for_brief and not pdf_text_for_brief:
            st.warning("⚠️ Brak danych wejściowych. Dodaj opis, komponenty lub PDF.")
        else:
            with st.spinner("Generuję opis zadania..."):
                try:
                    prompt_brief = build_brief_prompt(
                        st.session_state.get("description", ""),
                        components_for_brief + components_from_json_for_brief,
                        pdf_text_for_brief,
                        department
                    )

                    ai_model_brief = st.session_state.get("selected_text_model", "qwen2.5:7b")
                    resp = query_ollama(prompt_brief, model=ai_model_brief, format_json=True)
                    brief = parse_brief_response(resp)
                    st.session_state["ai_brief"] = brief
                    st.success("✅ Opis wygenerowany pomyślnie!")
                except Exception as e:
                    logger.exception("Brief generation failed")
                    st.error(f"❌ Nie udało się wygenerować opisu: {e}")
                    st.info("💡 Spróbuj ponownie lub zmień model AI w Sidebar")

    # Wyświetl brief (jeśli jest)
    if "ai_brief" in st.session_state:
        b = st.session_state["ai_brief"]
        if b.get("brief_md"):
            st.markdown(b["brief_md"])
        cols = st.columns(2)
        with cols[0]:
            if b.get("missing_info"):
                st.markdown("**Brakujące informacje / pytania:**")
                for it in b["missing_info"]:
                    st.write(f"• {it}")
            if b.get("assumptions"):
                st.markdown("**Założenia:**")
                for it in b["assumptions"]:
                    st.write(f"• {it}")
            if b.get("scope"):
                st.markdown("**Zakres:**")
                for it in b["scope"]:
                    st.write(f"• {it}")
        with cols[1]:
            if b.get("risks"):
                st.markdown("**Ryzyka:**")
                for r in b["risks"]:
                    st.write(f"• {r.get('risk','')} (impact: {r.get('impact','')})")
                    if r.get("mitigation"):
                        st.caption(f"Mitigation: {r['mitigation']}")
            if b.get("checklist"):
                st.markdown("**Checklist:**")
                for it in b["checklist"]:
                    st.write(f"☑︎ {it}")
            if b.get("open_questions"):
                st.markdown("**Otwarte pytania:**")
                for it in b["open_questions"]:
                    st.write(f"• {it}")

        # Pobranie do .md
        md_export = "# Opis zadania (AI)\n\n" + b.get("brief_md","") + "\n\n"
        if b.get("missing_info"):
            md_export += "## Brakujące informacje\n" + "\n".join([f"- {x}" for x in b["missing_info"]]) + "\n\n"
        if b.get("assumptions"):
            md_export += "## Założenia\n" + "\n".join([f"- {x}" for x in b["assumptions"]]) + "\n\n"
        if b.get("scope"):
            md_export += "## Zakres\n" + "\n".join([f"- {x}" for x in b["scope"]]) + "\n\n"
        if b.get("risks"):
            md_export += "## Ryzyka\n" + "\n".join([f"- {r['risk']} (impact: {r['impact']}) — {r.get('mitigation','')}" for r in b["risks"]]) + "\n\n"
        if b.get("checklist"):
            md_export += "## Checklist\n" + "\n".join([f"- {x}" for x in b["checklist"]]) + "\n\n"
        if b.get("open_questions"):
            md_export += "## Otwarte pytania\n" + "\n".join([f"- {x}" for x in b["open_questions"]]) + "\n\n"

        st.download_button("⬇️ Pobierz opis (.md)", md_export.encode("utf-8"),
                           file_name=f"opis_zadania_{datetime.now().strftime('%Y%m%d_%H%M')}.md",
                           mime="text/markdown")

    # Parametry sugestii
    st.subheader("⚙️ Uwzględnianie komentarzy")
    use_comments = st.checkbox("Uwzględnij sub‑komponenty z komentarzy w estymacji", value=True)
    conserv = st.slider("Konserwatywność proponowanych dodatków", min_value=0.5, max_value=1.5, value=1.0, step=0.1)
    enable_bundles = st.checkbox("Włącz podpowiedzi z historii (bundles)", value=True,
                                 help="Podpowiada typowe sub‑komponenty dla podobnych pozycji na bazie importów historycznych")

    # ═══════════════════════════════════════════════════════════
    # SPRINT 1: UI dla pytań doprecyzowujących
    # ═══════════════════════════════════════════════════════════
    st.subheader("💡 Pytania doprecyzowujące")

    # Sprawdź czy trzeba wygenerować pytania
    if (st.session_state.get("description") and
        len(st.session_state.get("description", "")) > 50 and
        not st.session_state.get("questions_answered")):

        # Zbierz PDF text dla kontekstu
        pdf_text_for_questions = ""
        if pdf_files:
            try:
                pdf_text_for_questions = "\n".join([extract_text_from_pdf(pf) for pf in pdf_files])
            except Exception:
                pass

        if st.session_state.get("pasted_text"):
            pdf_text_for_questions = (pdf_text_for_questions + "\n\n" + st.session_state.get("pasted_text", "")).strip()

        # Generuj pytania (tylko raz)
        if "clarifying_questions" not in st.session_state:
            with st.spinner("🤔 Generuję pytania doprecyzowujące..."):
                questions = generate_clarifying_questions(
                    st.session_state.get("description", ""),
                    department,
                    pdf_text_for_questions
                )
                st.session_state["clarifying_questions"] = questions

        questions = st.session_state.get("clarifying_questions", [])

        if questions and len(questions) > 0:
            st.info(f"🔍 Mam {len(questions)} pytań, które pomogą w dokładniejszej estymacji")

            with st.form("clarifying_questions_form"):
                st.markdown("### 📋 Pytania doprecyzowujące")
                st.caption("Odpowiedzi pomogą AI lepiej oszacować czas. Możesz pominąć pytania.")

                answers = {}
                for i, q in enumerate(questions):
                    st.markdown(f"**{i+1}. {q.get('question', '')}**")
                    if q.get('why'):
                        st.caption(f"💡 _{q['why']}_")

                    answer = st.text_area(
                        f"Odpowiedź {i+1}:",
                        key=f"answer_{i}",
                        placeholder="Zostaw puste jeśli nie wiesz lub nie dotyczy",
                        height=80
                    )
                    answers[q.get('question', '')] = answer
                    st.markdown("---")

                col_q1, col_q2 = st.columns(2)
                submit_questions = col_q1.form_submit_button("✅ Kontynuuj z odpowiedziami", type="primary")
                skip_questions = col_q2.form_submit_button("⏭️ Pomiń pytania")

                if submit_questions or skip_questions:
                    st.session_state["questions_answered"] = True
                    st.session_state["clarifying_answers"] = answers if submit_questions else {}
                    # Usuń pytania z sesji aby nie wygenerowały się ponownie
                    if "clarifying_questions" in st.session_state:
                        del st.session_state["clarifying_questions"]
                    st.rerun()

            # Zatrzymaj dalsze przetwarzanie dopóki nie odpowie
            st.warning("⏸️ Odpowiedz na pytania lub pomiń je, aby kontynuować analizę")
            return
        else:
            # Brak pytań - automatycznie oznacz jako odpowiedziane
            st.session_state["questions_answered"] = True
    elif not st.session_state.get("description") or len(st.session_state.get("description", "")) <= 50:
        st.info("💡 Dodaj opis projektu (min. 50 znaków), aby otrzymać pytania doprecyzowujące")
    else:
        if st.session_state.get("clarifying_answers"):
            st.success(f"✅ Odpowiedzi uwzględnione ({len([a for a in st.session_state['clarifying_answers'].values() if a])} odpowiedzi)")



    if st.button("🤖 Analizuj z AI", use_container_width=True):
        if not st.session_state.get("description") and not excel_file and not image_files and not pdf_files and not json_files and not pasted_text:
            st.warning("Podaj opis lub wgraj pliki")
        else:
            progress_bar = st.progress(0, text="Startuję...")
            try:
                # Excel
                components_from_excel = []
                if excel_file:
                    progress_bar.progress(15, text="Wczytuje Excel...")
                    components_from_excel = process_excel(excel_file)
    
                # JSON
                components_from_json = []
                if json_files:
                    progress_bar.progress(20, text="Czytam JSON...")
                    for jf in json_files:
                        try:
                            obj = safe_json_loads(jf.getvalue())
                            components_from_json += parse_components_from_docconv_json(obj)
                        except Exception:
                            pass
    
                # Obrazy
                images_b64 = []
                if image_files:
                    progress_bar.progress(25, text="Analizuję obrazy...")
                    for img in image_files:
                        images_b64.append(encode_image_b64(img))
    
                # PDF (+ tekst wklejony)
                pdf_text = ""
                if pdf_files:
                    progress_bar.progress(30, text="PDF...")
                    pdf_text = "\n".join([extract_text_from_pdf(pf) for pf in pdf_files])
                if pasted_text:
                    pdf_text = (pdf_text + "\n\n" + pasted_text).strip()
    
                # ════════════════════════════════════════════════════════════
                # KRYTYCZNE: Wzorce z DB
                # ════════════════════════════════════════════════════════════
                with get_db_connection() as conn, conn.cursor(cursor_factory=RealDictCursor) as cur:
                    cur.execute("""
                        SELECT name, avg_hours_total, avg_hours_3d_layout,
                               avg_hours_3d_detail, avg_hours_2d,
                               proportion_layout, proportion_detail, proportion_doc, occurrences
                        FROM component_patterns
                        WHERE department = %s AND occurrences > 2
                        ORDER BY occurrences DESC LIMIT 20
                    """, (department,))
                    learned_patterns = cur.fetchall()
    
                st.write(f"🧠 {len(learned_patterns)} wzorców z działu {department}")
    
                # Zbuduj komponenty dla promptu (okrojone do 30)
                components_for_prompt = (components_from_excel or []) + components_from_json

                # ════════════════════════════════════════════════════════════
                # SPRINT 1: Wzbogacenie opisu o odpowiedzi na pytania
                # ════════════════════════════════════════════════════════════
                enriched_description = st.session_state.get("description", "")
                if st.session_state.get("clarifying_answers"):
                    enriched_description += "\n\n--- DODATKOWE INFORMACJE ---\n"
                    for q, a in st.session_state["clarifying_answers"].items():
                        if a and a.strip():
                            enriched_description += f"\n{q}\nOdpowiedź: {a}\n"
                    st.info(f"✅ Wzbogacono opis o {len([a for a in st.session_state['clarifying_answers'].values() if a])} odpowiedzi")
                # ════════════════════════════════════════════════════════════

                # ════════════════════════════════════════════════════════════
                # KRYTYCZNE: UŻYJ build_analysis_prompt, NIE build_brief_prompt!
                # ════════════════════════════════════════════════════════════
                prompt = build_analysis_prompt(  # ✅✅✅ TO JEST POPRAWNE!
                    enriched_description,  # ⬅️ SPRINT 1: używaj wzbogaconego opisu
                    components_for_prompt,
                    learned_patterns,  # WAŻNE!
                    pdf_text,
                    department,
                    conn=conn  # ⬅️ DODAJ TO!

                )
                # ════════════════════════════════════════════════════════════
    
                # Wybór modelu
                if images_b64 and st.session_state.get("selected_vision_model"):
                    ai_model = st.session_state["selected_vision_model"]
                    st.info(f"🖼️ Używam modelu Vision: {ai_model}")
                else:
                    ai_model = st.session_state.get("selected_text_model", "qwen2.5:7b")
                    st.info(f"📝 Używam modelu tekstowego: {ai_model}")
    
                progress_bar.progress(60, text=f"AI ({ai_model})...")
                
                logger.info(f"🤖 Wysyłam prompt do AI ({ai_model}), długość: {len(prompt)} znaków")
                logger.debug(f"📝 Prompt preview: {prompt[:500]}...")
                ai_text = query_ollama(prompt, model=ai_model, images_b64=images_b64, format_json=True)

                # DEBUG - pokaż surową odpowiedź
                st.subheader("🔍 DEBUG: Surowa odpowiedź AI")
                st.code(ai_text, language="json")
                st.write(f"Długość: {len(ai_text)} znaków")
                
                logger.info(f"📥 Otrzymano odpowiedź AI, długość: {len(ai_text)} znaków")
                

                
                if not ai_text or len(ai_text) < 50:
                    st.error("❌ AI zwróciło pustą/zbyt krótką odpowiedź")
                    st.code(ai_text)
                    raise ValueError("Empty AI response")
    
                progress_bar.progress(80, text="Parsuję...")
                parsed = parse_ai_response(ai_text, components_from_excel=components_from_excel)
                # ═══════════════════════════════════════════════════════════
                # DODAJ TUTAJ DEBUG:
                st.subheader("🔍 DEBUG: Wynik parsowania")
                st.write("**parsed.get('components'):**")
                st.json(parsed.get('components', []))
                st.write(f"**Liczba komponentów:** {len(parsed.get('components', []))}")
                st.write(f"**total_hours:** {parsed.get('total_hours', 0)}")
                st.write(f"**Warnings:** {parsed.get('warnings', [])}")
                # ═══════════════════════════════════════════════════════════
                
                logger.info(f"📊 Sparsowano: {len(parsed.get('components',[]))} komponentów, total: {parsed.get('total_hours',0):.1f}h")
                if parsed.get('warnings'):
                    logger.warning(f"⚠️ Warningi parsera: {parsed['warnings']}")

            # ... reszta kodu (kategoryzacja, zapisanie do session_state, itd.)
#---------------------------------------------------------+++++++


                
            # ════════════════════════════════════════════════════════════=================================
                # 🌐 Web enhancement (opcjonalne - po parsowaniu)
                if st.session_state.get("allow_web_lookup") and parsed.get('components'):
                    progress_bar.progress(85, text="🌐 Wzbogacam o dane z sieci...")
                    enhanced_count = 0
                    for comp in parsed['components'][:5]:  # Tylko pierwsze 5
                        try:
                            web_data = enhance_estimation_with_web(
                                comp.get('name', ''), 
                                department, 
                                enable_web=True
                            )
                            if web_data.get("web_context"):
                                comp["web_notes"] = web_data["web_context"]
                                enhanced_count += 1
                        except Exception as e:
                            logger.warning(f"Web enhancement failed for '{comp.get('name')}': {e}")
                    
                    if enhanced_count > 0:
                        st.info(f"✅ Wzbogacono {enhanced_count} komponentów danymi z sieci")

                progress_bar.progress(90, text="Finalizuję...")

                # Dołącz komponenty z JSON (deduplikacja po canonicalize_name)
                if components_from_json:
                    parsed['components'] = merge_components(parsed.get('components', []), components_from_json)
                    parsed['total_layout'] = sum(c.get('hours_3d_layout', 0) for c in parsed['components'])
                    parsed['total_detail'] = sum(c.get('hours_3d_detail', 0) for c in parsed['components'])
                    parsed['total_2d'] = sum(c.get('hours_2d', 0) for c in parsed['components'])

                # Kategoryzacja
                if parsed.get('components'):
                    for c in parsed['components']:
                        if not c.get('is_summary', False):
                            c['category'] = categorize_component(c.get('name', ''))

                st.session_state["ai_analysis"] = parsed
                st.session_state["base_components"] = parsed.get('components', [])
                st.session_state["ai_adjustments"] = parsed.get('ai_adjustments', [])

                # Propozycje z komentarzy (patterns/heurystyki)
                if use_comments:
                    with get_db_connection() as conn:
                        proposals = propose_adjustments_for_components(conn, st.session_state["base_components"], department, conserv)
                    st.session_state["rule_adjustments"] = proposals
                else:
                    st.session_state["rule_adjustments"] = []

                progress_bar.progress(100, text="Gotowe ✅")
                time.sleep(0.6)
                progress_bar.empty()

            except Exception as e:
                logger.exception("Analiza failed")
                st.error(f"Błąd: {e}")

    # Wyniki analizy i edycja
    if "ai_analysis" in st.session_state:
        analysis = st.session_state["ai_analysis"]
        base_components = st.session_state.get("base_components", [])
        ai_adjustments = st.session_state.get("ai_adjustments", [])
        rule_adjustments = st.session_state.get("rule_adjustments", [])

        # Bundles z historii dla komponentów bez sub‑komponentów
        bundle_adjustments = []
        if enable_bundles and base_components:
            with get_db_connection() as conn:
                for comp in base_components:
                    if comp.get('is_summary') or comp.get('subcomponents'):
                        continue
                    adds = propose_bundles_for_component(conn, comp.get('name',''), department, conservativeness=conserv)
                    if adds:
                        bundle_adjustments.append({"parent": comp.get('name',''), "adds": adds})

        combined_rule_adjustments = (rule_adjustments or []) + (bundle_adjustments or [])

        st.subheader("Wynik analizy")
        for w in analysis.get("warnings", []):
            st.warning(w)

        layout_h = analysis.get("total_layout", 0.0)
        detail_h = analysis.get("total_detail", 0.0)
        doc_h = analysis.get("total_2d", 0.0)
        estimated_hours = max(0.0, layout_h + detail_h + doc_h)

        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Layout", f"{layout_h:.1f}h")
        col2.metric("Detail", f"{detail_h:.1f}h")
        col3.metric("2D", f"{doc_h:.1f}h")
        hourly_rate = st.sidebar.number_input("Stawka PLN/h", min_value=1, max_value=1000, value=150, step=10)
        col4.metric("TOTAL", f"{estimated_hours:.1f}h", delta=f"{(estimated_hours * hourly_rate):.0f} PLN")

        # Proponowane dodatki (AI)
        st.subheader("💡 Proponowane dodatki (AI z komentarzy)")
        ai_selected = []
        if ai_adjustments:
            for i, adj in enumerate(ai_adjustments):
                parent = adj.get("parent", "komponent")
                with st.expander(f"AI: {parent}"):
                    for j, add in enumerate(adj.get("adds", [])):
                        key = f"ai_adj_{i}_{j}"
                        default = True
                        checked = st.checkbox(
                            f"{add['qty']}x {add['name']} → +L {add['layout_add']:.1f}h, +D {add['detail_add']:.1f}h, +2D {add['doc_add']:.1f}h",
                            value=default, key=key
                        )
                        st.caption(f"Powód: {add.get('reason','')}")
                        if checked:
                            ai_selected.append({"parent": parent, "add": add})
        else:
            st.caption("Brak propozycji AI lub model nie zwrócił 'adjustments'.")

        # Proponowane dodatki (Wzorce/Heurystyki/Historia)
        st.subheader("🧠 Proponowane dodatki (wzorce/heurystyki + historia bundles)")
        rule_selected = []
        if combined_rule_adjustments:
            for i, adj in enumerate(combined_rule_adjustments):
                parent = adj.get("parent", "komponent")
                with st.expander(f"Wzorce/Heurystyki/Historia: {parent}"):
                    for j, add in enumerate(adj.get("adds", [])):
                        key = f"rule_adj_{i}_{j}"
                        default = True if add.get("source") == "pattern" else False
                        checked = st.checkbox(
                            f"{add['qty']}x {add['name']} → +L {add['layout_add']:.1f}h, +D {add['detail_add']:.1f}h, +2D {add['doc_add']:.1f}h  ({add.get('source','')}, conf={add.get('confidence',0):.2f})",
                            value=default, key=key
                        )
                        st.caption(f"Powód: {add.get('reason','')}")
                        if checked:
                            rule_selected.append({"parent": parent, "add": add})
        else:
            st.caption("Brak propozycji z komentarzy/historii lub funkcja wyłączona.")

        # Zbuduj komponenty z zaakceptowanych dodatków (AI + reguły)
        adjustment_components = []
        for src, group in [("AI", ai_selected), ("RULE", rule_selected)]:
            for item in group:
                parent = item["parent"]
                add = item["add"]
                comp = {
                    "name": f"ADJ: {add['name']} (x{add['qty']})",
                    "hours_3d_layout": float(add["layout_add"]),
                    "hours_3d_detail": float(add["detail_add"]),
                    "hours_2d": float(add["doc_add"]),
                    "hours": float(add["layout_add"] + add["detail_add"] + add["doc_add"]),
                    "is_adjustment": True,
                    "parent": parent,
                    "source": src
                }
                adjustment_components.append(comp)

        st.subheader("🔧 Edytuj wycenę (komponenty bazowe)")
        final_components_base = []
        if base_components:
            parts_only = [
                c for c in base_components
                if not c.get('is_summary', False) and c.get('name') not in ['[part]', '[assembly]', '', ' ']
            ]
            st.caption(f"ℹ️ Pokazano {len(parts_only)} komponentów")
            for i, comp in enumerate(parts_only):
                display_name = comp['name'][:50] + "..." if len(comp['name']) > 50 else comp['name']
                with st.expander(f"{display_name} - {comp.get('hours', 0):.1f}h"):
                    st.markdown(f"**Pełna nazwa:** {comp['name']}")
                    c1, c2, c3 = st.columns(3)
                    new_layout = c1.number_input("Layout", value=float(comp.get('hours_3d_layout', 0)), key=f"l_{i}")
                    new_detail = c2.number_input("Detail", value=float(comp.get('hours_3d_detail', 0)), key=f"d_{i}")
                    new_doc = c3.number_input("2D", value=float(comp.get('hours_2d', 0)), key=f"doc_{i}")
                    comp2 = dict(comp)
                    comp2['hours_3d_layout'] = new_layout
                    comp2['hours_3d_detail'] = new_detail
                    comp2['hours_2d'] = new_doc
                    comp2['hours'] = new_layout + new_detail + new_doc
                    final_components_base.append(comp2)
                    if comp.get('subcomponents'):
                        st.markdown("**Zawiera:**")
                        for sub in comp['subcomponents']:
                            qty = sub.get('quantity', 1)
                            st.text(f"  • {qty}x {sub['name']}" if qty > 1 else f"  • {sub['name']}")
                    if comp.get('comment'):
                        st.caption(f"💬 {comp['comment']}")

        # Połączenie bazowych i dodatków
        combined_components = final_components_base + adjustment_components

        # Współczynniki z Excela
        if "excel_multipliers" in st.session_state and combined_components:
            st.subheader("📊 Współczynniki z Excela")
            mult = st.session_state["excel_multipliers"]
            c1, c2, c3 = st.columns(3)
            c1.metric("Layout", f"x{mult['layout']:.2f}")
            c2.metric("Detail", f"x{mult['detail']:.2f}")
            c3.metric("Doc", f"x{mult['documentation']:.2f}")
            apply_mult = st.checkbox("Zastosuj współczynniki do wszystkich pozycji (w tym dodatków)", value=False, key="apply_mult")
            if apply_mult:
                combined_scaled = []
                for c in combined_components:
                    c2 = dict(c)
                    c2['hours_3d_layout'] = c.get('hours_3d_layout', 0) * mult['layout']
                    c2['hours_3d_detail'] = c.get('hours_3d_detail', 0) * mult['detail']
                    c2['hours_2d'] = c.get('hours_2d', 0) * mult['documentation']
                    c2['hours'] = c2['hours_3d_layout'] + c2['hours_3d_detail'] + c2['hours_2d']
                    combined_scaled.append(c2)
                combined_components = combined_scaled

        # Podsumowanie końcowe
        sum_layout = sum(c.get('hours_3d_layout', 0) for c in combined_components if not c.get('is_summary', False))
        sum_detail = sum(c.get('hours_3d_detail', 0) for c in combined_components if not c.get('is_summary', False))
        sum_doc = sum(c.get('hours_2d', 0) for c in combined_components if not c.get('is_summary', False))
        sum_total = sum_layout + sum_detail + sum_doc

        st.metric("🔢 Suma (po dodatkach i multipliers)", f"{sum_total:.1f}h")

        st.subheader("🗂️ Harmonogram")
        show_project_timeline(combined_components)

        # Podobne projekty (keyword i semantycznie)
        with get_db_connection() as conn:
            similar = find_similar_projects(conn, st.session_state.get("description"), department)
        st.subheader(f"📊 Podobne projekty ({department})")
        if similar:
            for proj in similar:
                cc1, cc2, cc3 = st.columns([3,1,1])
                cc1.write(f"**{proj['name']}** ({proj['client'] or '-'})")
                cc2.metric("Szacowano", f"{(proj['estimated_hours'] or 0):.1f}h")
                if proj['actual_hours']:
                    cc3.metric("Rzeczywiście", f"{proj['actual_hours']:.1f}h")
        else:
            st.info("Brak podobnych")

        with get_db_connection() as conn:
            similar_sem = find_similar_projects_semantic(conn, st.session_state.get("description"), department)
        st.subheader(f"🧭 Semantycznie podobne projekty (pgvector)")
        if similar_sem:
            for sp in similar_sem:
                sim_pct = sp['similarity'] * 100
                st.write(f"- **{sp['name']}** (sim={sim_pct:.0f}%) — est: {(sp['estimated_hours'] or 0):.1f}h" +
                         (f", act: {sp['actual_hours']:.1f}h" if sp['actual_hours'] else ""))
        else:
            st.caption("Brak embeddingów — dodaj projekty i uruchom przeliczanie")

        # Eksport do Excel
        st.subheader("📤 Eksport")
        if st.button("📥 Export do Excel"):
            excel_data = export_quotation_to_excel({
                'name': st.session_state.get("project_name"),
                'client': st.session_state.get("client"),
                'department': department,
                'components': combined_components,
                'total_hours': sum_total
            })
            st.download_button("⬇️ Pobierz", excel_data,
                               file_name=f"wycena_{st.session_state.get('project_name','p')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # Zapis projektu
        st.subheader("💾 Zapisz projekt")
        c1, c2 = st.columns([3,1])
        with c1:
            change_desc = st.text_input("Opis zmian", placeholder="np. 'Pierwsza wycena'")
        with c2:
            is_approved = st.checkbox("Zatwierdzone", value=False)

        if st.button("💾 Zapisz", type="primary", use_container_width=True):
            errors = []
            name = st.session_state.get("project_name")
            if not name or not name.strip():
                errors.append("Nazwa nie może być pusta")
            if sum_total < 0:
                errors.append("Godziny < 0")
            if errors:
                for e in errors:
                    st.error(e)
            else:
                try:
                    components_to_save = [c for c in combined_components if not c.get('is_summary', False)]
                    with get_db_connection() as conn, conn.cursor() as cur:
                        cur.execute("SELECT id FROM projects WHERE name = %s AND department = %s",
                                    (name, department))
                        existing = cur.fetchone()
                        if existing:
                            project_id = existing[0]
                            cur.execute("SELECT COUNT(*) FROM project_versions WHERE project_id = %s", (project_id,))
                            version_num = f"v1.{cur.fetchone()[0] + 1}"
                            cur.execute("""
                                UPDATE projects SET components = %s,
                                estimated_hours_3d_layout = %s, estimated_hours_3d_detail = %s,
                                estimated_hours_2d = %s, estimated_hours = %s,
                                ai_analysis = %s, updated_at = NOW()
                                WHERE id = %s
                            """, (json.dumps(components_to_save, ensure_ascii=False),
                                  float(sum_layout), float(sum_detail),
                                  float(sum_doc), float(sum_total),
                                  analysis["raw_text"], project_id))

                            save_project_version(conn, project_id, version_num, components_to_save,
                                                sum_total, sum_layout, sum_detail, sum_doc,
                                                change_desc or "", "System")

                            if is_approved:
                                cur.execute("UPDATE project_versions SET is_approved = TRUE WHERE project_id = %s AND version = %s",
                                          (project_id, version_num))
                            conn.commit()
                            st.success(f"✅ Zaktualizowano! {version_num}")
                        else:
                            cur.execute("""
                                INSERT INTO projects (name, client, department, description, components,
                                estimated_hours_3d_layout, estimated_hours_3d_detail, estimated_hours_2d,
                                estimated_hours, ai_analysis)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id
                            """, (
                                (st.session_state.get("project_name") or "").strip(),
                                (st.session_state.get("client") or "").strip(),
                                department, (st.session_state.get("description") or "").strip(),
                                json.dumps(components_to_save, ensure_ascii=False),
                                float(sum_layout), float(sum_detail),
                                float(sum_doc), float(sum_total),
                                analysis["raw_text"]
                            ))
                            project_id = cur.fetchone()[0]
                            ensure_project_embedding(cur, project_id, st.session_state.get("description", ""))

                            save_project_version(conn, project_id, "v1.0", components_to_save,
                                                sum_total, sum_layout, sum_detail, sum_doc,
                                                change_desc or "Pierwsza wycena", "System")

                            if is_approved:
                                cur.execute("UPDATE project_versions SET is_approved = TRUE WHERE project_id = %s AND version = 'v1.0'",
                                          (project_id,))
                            conn.commit()
                            st.success(f"✅ Zapisano! ID: {project_id}")

                        logger.info(f"Zapisano: {project_id} - {st.session_state.get('project_name')}")
                        st.balloons()
                        time.sleep(1.0)
                        st.rerun()
                except Exception as e:
                    st.error(f"Błąd: {e}")
                    logger.exception("Zapis failed")
# === CAD Estimator Pro — main.py (Part 4/4) ==================================
# Dashboard, Historia i Uczenie, Generatory demo, Sidebar i main()
# ============================================================================

# === GENERATORY DEMO ===
def generate_sample_excel() -> bytes:
    """
    Generuje przykładowy Excel pasujący do parsera:
    - Multipliers w wierszu 10 (index 9): kolumny H, J, L (7,9,11)
    - Dane od wiersza 12 (index 11)
    """
    output = BytesIO()
    try:
        # preferuj xlsxwriter (jeśli dostępny)
        import xlsxwriter  # noqa
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            pd.DataFrame().to_excel(writer, sheet_name='Dane', index=False)
            ws = writer.sheets['Dane']

            # Multipliers (wiersz 10 zero-based -> index 9)
            ws.write(9, 7, 1.0)   # Layout
            ws.write(9, 9, 1.0)   # Detail
            ws.write(9, 11, 1.0)  # Doc

            # Nagłówki (opcjonalnie, wiersz 11 -> index 10)
            headers = ["Pozycja", "Opis", "Komentarz", "Części std", "Części spec", "", "", "Layout [h]", "", "Detail [h]", "", "Doc [h]"]
            for col, h in enumerate(headers):
                ws.write(10, col, h)

            # Dane od wiersza 12 (index 11)
            row = 11
            ws.write(row, 0, "1,0"); ws.write(row, 1, "Stacja dociskania omega (złożenie)"); row += 1

            ws.write(row, 0, "1,1")
            ws.write(row, 1, "Dociski omega boczna; blachy")
            ws.write(row, 2, "2x - docisk śrubowy odrzucany; śruba trapezowa; konsola docisku")
            ws.write(row, 7, 2.0); ws.write(row, 9, 6.0); ws.write(row, 11, 3.0); row += 1

            ws.write(row, 0, "1,2")
            ws.write(row, 1, "Konsola główna")
            ws.write(row, 2, "płyta montażowa; 4x wspornik; osłona boczna")
            ws.write(row, 7, 1.0); ws.write(row, 9, 4.0); ws.write(row, 11, 2.0); row += 1

            ws.write(row, 0, "1,3")
            ws.write(row, 1, "Płyta bazowa z otworami")
            ws.write(row, 2, "8x otwór M12; fazowanie")
            ws.write(row, 7, 0.5); ws.write(row, 9, 2.5); ws.write(row, 11, 1.0)
    except Exception:
        # fallback: openpyxl (minimalny)
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df = pd.DataFrame([
                ["1,0", "Stacja dociskania omega (złożenie)", None, None, None, None, None, None, None, None, None, None],
                ["1,1", "Dociski omega boczna; blachy", "2x - docisk śrubowy odrzucany; śruba trapezowa; konsola docisku",
                 0, 0, None, None, 2.0, None, 6.0, None, 3.0],
                ["1,2", "Konsola główna", "płyta montażowa; 4x wspornik; osłona boczna",
                 0, 0, None, None, 1.0, None, 4.0, None, 2.0],
                ["1,3", "Płyta bazowa z otworami", "8x otwór M12; fazowanie",
                 0, 0, None, None, 0.5, None, 2.5, None, 1.0],
            ])
            df.to_excel(writer, sheet_name='Dane', header=False, index=False)
    output.seek(0)
    return output.getvalue()

def generate_sample_pdf() -> bytes | None:
    """Generuje prosty PDF (wymaga reportlab). Jeśli brak, zwraca None."""
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        buf = BytesIO()
        c = canvas.Canvas(buf, pagesize=A4)
        textobject = c.beginText(40, 800)
        lines = [
            "Specyfikacja: Stacja dociskania omega (boczna)",
            "- Wymagane dociski śrubowe z mechanizmem odrzucania",
            "- Konsola docisku i płyta bazowa",
            "- Śruby trapezowe w mechanizmie odrzutu",
            "Normy: ISO 12100, EN 1090",
            "Uwagi: kinematyka docisku, docisk boczny, kontrola luzu"
        ]
        for l in lines:
            textobject.textLine(l)
        c.drawText(textobject)
        c.showPage()
        c.save()
        buf.seek(0)
        return buf.getvalue()
    except Exception as e:
        logger.warning(f"Nie można wygenerować PDF (brak reportlab?): {e}")
        return None

def generate_sample_image() -> bytes:
    """Generuje prosty obraz PNG (schemat poglądowy) do testu."""
    w, h = 800, 400
    img = Image.new("RGB", (w, h), (245, 245, 245))
    draw = ImageDraw.Draw(img)
    draw.rectangle([20, 20, w-20, h-20], outline=(50, 50, 50), width=3)
    draw.rectangle([60, 150, 220, 250], outline="navy", width=3); draw.text((70, 260), "Płyta bazowa", fill="navy")
    draw.rectangle([300, 120, 520, 180], outline="darkgreen", width=3); draw.text((310, 185), "Docisk śrubowy", fill="darkgreen")
    draw.line([520, 150, 700, 150], fill="black", width=3); draw.text((600, 160), "Odrzut", fill="black")
    draw.text((30, 30), "Stacja dociskania omega (schemat poglądowy)", fill=(0, 0, 0))
    buf = BytesIO(); img.save(buf, format='PNG'); buf.seek(0)
    return buf.getvalue()

def fill_demo_fields():
    st.session_state['project_name'] = "Stacja dociskania omega - DEMO"
    st.session_state['client'] = "Klient Demo sp. z o.o."
    st.session_state['description'] = (
        "Stacja dociskania detalu typu omega z dociskami bocznymi. "
        "Wymagania: mechanizm odrzutu docisku śrubowego, konsola docisku, płyta bazowa. "
        "Normy: ISO 12100, EN 1090. Złożoność średnia, kinematyka docisków."
    )
    st.success("Wypełniono formularz przykładowymi danymi.")

# === UI: Dashboard ===
def render_dashboard_page():
    st.header("📊 Dashboard")
    with get_db_connection() as conn, conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute("SELECT COUNT(*) as count FROM projects")
        project_count = cur.fetchone()['count']
        cur.execute("SELECT AVG(accuracy) as avg FROM projects WHERE accuracy IS NOT NULL")
        avg_accuracy = (cur.fetchone() or {}).get('avg') or 0
        cur.execute("""
        SELECT department, COUNT(*) as count
        FROM projects WHERE department IS NOT NULL
        GROUP BY department ORDER BY department
        """)
        dept_stats = cur.fetchall()

    c1, c2 = st.columns(2)
    c1.metric("Projekty", project_count)
    c2.metric("Średnia dokładność", f"{avg_accuracy*100:.1f}%")

    if dept_stats:
        st.subheader("Projekty wg działów")
        df_dept = pd.DataFrame(dept_stats)
        df_dept['department_name'] = df_dept['department'].map(DEPARTMENTS)
        st.bar_chart(df_dept.set_index('department_name')['count'])

    st.header("🔍 Wyszukaj projekty")
    search_dept = st.selectbox("Dział", options=[''] + list(DEPARTMENTS.keys()),
                               format_func=lambda x: 'Wszystkie' if x == '' else f"{x} - {DEPARTMENTS[x]}")
    search_query = st.text_input("Słowa kluczowe")
    if search_query:
        with get_db_connection() as conn, conn.cursor(cursor_factory=RealDictCursor) as cur:
            if search_dept:
                cur.execute("""
                    SELECT id, name, client, department, estimated_hours, description
                    FROM projects WHERE department = %s
                    AND to_tsvector('simple', coalesce(name,'') || ' ' || coalesce(client,'') || ' ' || coalesce(description,'')) 
                        @@ websearch_to_tsquery('simple', %s)
                    ORDER BY created_at DESC LIMIT 10
                """, (search_dept, search_query))
            else:
                cur.execute("""
                    SELECT id, name, client, department, estimated_hours, description
                    FROM projects
                    WHERE to_tsvector('simple', coalesce(name,'') || ' ' || coalesce(client,'') || ' ' || coalesce(description,'')) 
                        @@ websearch_to_tsquery('simple', %s)
                    ORDER BY created_at DESC LIMIT 10
                """, (search_query,))
            results = cur.fetchall()
        if results:
            st.write(f"Znaleziono {len(results)} projektów:")
            df_results = pd.DataFrame(results)
            df_results['department_name'] = df_results['department'].map(DEPARTMENTS)
            st.dataframe(df_results, use_container_width=True)

            selected_project = st.selectbox(
                "Historia wersji",
                options=results,
                format_func=lambda p: f"{p['name']} ({p['department']})"
            )
            if selected_project:
                with get_db_connection() as conn:
                    versions = get_project_versions(conn, selected_project['id'])
                if versions:
                    st.subheader(f"📜 Historia: {selected_project['name']}")
                    for v in versions:
                        with st.expander(f"{v['version']} - {v['created_at'].strftime('%Y-%m-%d %H:%M')} {'✅' if v['is_approved'] else ''}",
                                         expanded=(v == versions[0])):
                            col1, col2, col3 = st.columns(3)
                            col1.metric("Layout", f"{v['estimated_hours_3d_layout']:.1f}h")
                            col2.metric("Detail", f"{v['estimated_hours_3d_detail']:.1f}h")
                            col3.metric("2D", f"{v['estimated_hours_2d']:.1f}h")
                            st.metric("TOTAL", f"{v['estimated_hours']:.1f}h")
                            if v['change_description']:
                                st.text_area("Opis", v['change_description'], height=100, disabled=True, key=f"d_{v['id']}")
                            st.caption(f"Autor: {v['changed_by']}")
        else:
            st.info("Nie znaleziono")

# === UI: Historia i Uczenie ===
def render_history_page():
    st.header("📚 Historia i Uczenie")
    tab1, tab2, tab3 = st.tabs(["✏️ Feedback", "🧠 Wzorce", "📦 Batch Import"])

    # === TAB 1: Feedback (rzeczywiste godziny → uczenie wzorców) ===
    with tab1:
        st.subheader("Dodaj feedback")
        feedback_dept = st.selectbox("Dział", options=[''] + list(DEPARTMENTS.keys()),
                                     format_func=lambda x: 'Wszystkie' if x == '' else f"{x} - {DEPARTMENTS[x]}")

        with get_db_connection() as conn, conn.cursor(cursor_factory=RealDictCursor) as cur:
            if feedback_dept:
                cur.execute("""
                    SELECT id, name, department, estimated_hours
                    FROM projects WHERE actual_hours IS NULL AND department = %s
                    ORDER BY created_at DESC
                """, (feedback_dept,))
            else:
                cur.execute("""
                    SELECT id, name, department, estimated_hours 
                    FROM projects 
                    WHERE actual_hours IS NULL 
                    ORDER BY created_at DESC
                """)
            pending = cur.fetchall()

        if pending:
            proj = st.selectbox("Projekt", options=pending,
                                format_func=lambda p: f"[{p['department']}] {p['name']} (ID: {p['id']}) | est: {p['estimated_hours']:.1f}h")
            actual_hours = st.number_input("Rzeczywiste godziny", min_value=0.0, step=0.5, value=float(proj['estimated_hours']))

            if st.button("💾 Zapisz feedback", type="primary"):
                if actual_hours <= 0:
                    st.error("Godziny > 0")
                else:
                    with get_db_connection() as conn, conn.cursor() as cur:
                        estimated = float(proj['estimated_hours'])
                        accuracy = 1 - abs(estimated - actual_hours) / estimated if estimated > 0 else 0

                        cur.execute("UPDATE projects SET actual_hours = %s, accuracy = %s WHERE id = %s",
                                    (actual_hours, accuracy, proj['id']))

                        cur.execute("SELECT components, department FROM projects WHERE id = %s", (proj['id'],))
                        row = cur.fetchone()
                        components_data = (row[0] or [])
                        dept = row[1]

                        if components_data:
                            ratio = actual_hours / estimated if estimated > 0 else 1.0
                            # ucz wzorce komp. głównych + subów
                            for comp in components_data:
                                if comp.get('is_summary'):
                                    continue
                                layout_est = float(comp.get('hours_3d_layout', 0))
                                detail_est = float(comp.get('hours_3d_detail', 0))
                                doc_est = float(comp.get('hours_2d', 0))
                                total_est = float(comp.get('hours', 0))
                                if total_est > 0:
                                    update_pattern_smart(
                                        cur, comp.get('name', 'nieznany'), dept,
                                        layout_est * ratio, detail_est * ratio, doc_est * ratio, source='actual'
                                    )
                                    subs = comp.get('subcomponents', [])
                                    if subs:
                                        total_qty = sum(s.get('quantity', 1) for s in subs)
                                        for sub in subs:
                                            qty = sub.get('quantity', 1)
                                            weight = qty / total_qty if total_qty > 0 else 1.0 / len(subs)
                                            sub_layout = layout_est * ratio * weight
                                            sub_detail = detail_est * ratio * weight
                                            sub_doc = doc_est * ratio * weight
                                            update_pattern_smart(cur, sub['name'], dept, sub_layout, sub_detail, sub_doc, source='subcomponent')

                            # baseline kategorii
                            agg_cat = {}
                            for comp in components_data:
                                if comp.get('is_summary'):
                                    continue
                                layout_act = float(comp.get('hours_3d_layout', 0)) * ratio
                                detail_act = float(comp.get('hours_3d_detail', 0)) * ratio
                                doc_act = float(comp.get('hours_2d', 0)) * ratio
                                cat = comp.get('category') or categorize_component(comp.get('name',''))
                                agg_cat.setdefault(cat, [0.0, 0.0, 0.0])
                                agg_cat[cat][0] += layout_act; agg_cat[cat][1] += detail_act; agg_cat[cat][2] += doc_act
                            for cat, (l, d, dc) in agg_cat.items():
                                update_category_baseline(cur, dept, cat, l, d, dc)

                        conn.commit()
                    st.success("Dzięki! System się zaktualizował.")
                    time.sleep(1.0)
                    st.rerun()
        else:
            st.info("🎉 Wszystkie projekty mają feedback!")

    # === TAB 2: Wzorce komponentów (podgląd + narzędzia admina) ===
    with tab2:
        st.subheader("Wzorce komponentów")
        pattern_dept = st.selectbox("Filtruj", options=[''] + list(DEPARTMENTS.keys()),
                                    format_func=lambda x: 'Wszystkie' if x == '' else f"{x} - {DEPARTMENTS[x]}")

        with get_db_connection() as conn, conn.cursor(cursor_factory=RealDictCursor) as cur:
            if pattern_dept:
                cur.execute("""
                    SELECT name, department, avg_hours_total, avg_hours_3d_layout,
                           avg_hours_3d_detail, avg_hours_2d, proportion_layout,
                           proportion_detail, proportion_doc, occurrences
                    FROM component_patterns
                    WHERE department = %s AND occurrences > 0 ORDER BY occurrences DESC
                """, (pattern_dept,))
            else:
                cur.execute("""
                    SELECT name, department, avg_hours_total, avg_hours_3d_layout,
                           avg_hours_3d_detail, avg_hours_2d, proportion_layout,
                           proportion_detail, proportion_doc, occurrences
                    FROM component_patterns WHERE occurrences > 0
                    ORDER BY department, occurrences DESC
                """)
            patterns = cur.fetchall()

        if patterns:
            df = pd.DataFrame(patterns)
            df['department_name'] = df['department'].map(DEPARTMENTS)
            df['proportion_layout'] = (df['proportion_layout'] * 100).round(1).astype(str) + '%'
            df['proportion_detail'] = (df['proportion_detail'] * 100).round(1).astype(str) + '%'
            df['proportion_doc'] = (df['proportion_doc'] * 100).round(1).astype(str) + '%'
            st.dataframe(df, use_container_width=True)
            st.info(f"{len(patterns)} wzorców")
        else:
            st.info("Brak wzorców")

        with st.expander("🧰 Admin: przelicz embeddingi"):
            if st.button("🔄 Przelicz embeddingi dla istniejących danych"):
                with get_db_connection() as conn, conn.cursor(cursor_factory=RealDictCursor) as cur:
                    cur.execute("SELECT id, description FROM projects WHERE description IS NOT NULL")
                    projects_to_embed = cur.fetchall()
                    cur.execute("SELECT pattern_key, department, name FROM component_patterns WHERE pattern_key IS NOT NULL")
                    patterns_to_embed = cur.fetchall()
                    total_items = len(projects_to_embed) + len(patterns_to_embed)
                    if total_items == 0:
                        st.warning("Brak danych do przeliczenia")
                    else:
                        progress = st.progress(0, text="Przeliczam embeddingi...")
                        for idx, p in enumerate(projects_to_embed):
                            ensure_project_embedding(cur, p['id'], p['description'])
                            progress.progress(int((idx + 1) / total_items * 100), text=f"Projekty: {idx+1}/{len(projects_to_embed)}")
                        for idx, r in enumerate(patterns_to_embed):
                            ensure_pattern_embedding(cur, r['pattern_key'], r['department'], r['name'])
                            progress.progress(int((len(projects_to_embed) + idx + 1) / total_items * 100), text=f"Wzorce: {idx+1}/{len(patterns_to_embed)}")
                        conn.commit()
                        progress.empty()
                        st.success(f"✅ Przeliczono {len(projects_to_embed)} projektów + {len(patterns_to_embed)} wzorców")

    # === TAB 3: Batch Import (A1 → opis) + Edycja po imporcie ===
    with tab3:
        st.subheader("📦 Batch Import (opis z A1 pierwszej zakładki)")
        st.info("Podczas importu opis projektu zostanie automatycznie pobrany z komórki A1 pierwszego arkusza. "
                "Jeśli A1 jest puste, zapisze się placeholder 'Projekt historyczny: <nazwa>'. "
                "Po imporcie możesz opisy edytować niżej.")

        batch_dept = st.selectbox("Dział dla importu", options=list(DEPARTMENTS.keys()),
                                  format_func=lambda x: f"{x} - {DEPARTMENTS[x]}", key="batch_dept")

        excel_files = st.file_uploader("Excel (wiele)", type=['xlsx', 'xls'], accept_multiple_files=True, key="batch")
        if excel_files:
            st.write(f"📁 {len(excel_files)} plików")
            for f in excel_files[:10]:
                st.write(f"• {f.name}")
            if len(excel_files) > 10:
                st.write(f"... +{len(excel_files) - 10}")

            learn_from_import = st.checkbox("Ucz wzorce z importu (komponenty + sub‑komponenty)", value=True)
            distribute_method = st.radio(
                "Rozdział godzin na sub‑komponenty",
                options=['qty', 'equal'],
                format_func=lambda v: "Proporcjonalnie do ilości (qty)" if v == 'qty' else "Po równo",
                horizontal=True
            )

            if st.button("🚀 Importuj", type="primary", use_container_width=True):
                st.info(f"Import {len(excel_files)} do {batch_dept}...")
                results = batch_import_excels(excel_files, batch_dept,
                                              learn_from_import=learn_from_import,
                                              distribute=distribute_method)
                success = sum(1 for r in results if r['status'] == 'success')
                errors = sum(1 for r in results if r['status'] == 'error')
                c1, c2 = st.columns(2)
                c1.metric("✅ Sukces", success)
                c2.metric("❌ Błędy", errors)
                st.subheader("Szczegóły")
                df = pd.DataFrame(results)
                st.dataframe(df, use_container_width=True)
                if success > 0:
                    st.success(f"🎉 {success} projektów!")
                if errors > 0:
                    st.warning(f"⚠️ {errors} błędów")

        st.markdown("---")
        st.subheader("✏️ Uzupełnij/edytuj opisy po imporcie")
        st.caption("Poniżej widzisz ostatnie projekty historyczne bez opisu lub z opisem zastępczym. Uzupełnij i zapisz.")

        with get_db_connection() as conn, conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT id, name, department, coalesce(description,'') AS description
                FROM projects
                WHERE is_historical = TRUE
                  AND (description IS NULL OR trim(description) = '' OR description LIKE 'Projekt historyczny:%')
                ORDER BY created_at DESC
                LIMIT 50
            """)
            missing = cur.fetchall()

        if missing:
            for p in missing:
                st.markdown(f"**[{p['department']}] {p['name']}**")
                new_desc = st.text_area("Opis", value=p['description'], key=f"d_{p['id']}", height=100)
                save_col, _ = st.columns([1,3])
                if save_col.button("💾 Zapisz opis", key=f"save_{p['id']}"):
                    with get_db_connection() as conn, conn.cursor() as cur:
                        cur.execute("UPDATE projects SET description=%s WHERE id=%s", (new_desc.strip(), p['id']))
                        ensure_project_embedding(cur, p['id'], new_desc.strip())
                        conn.commit()
                    st.success("Zapisano opis ✔️")
                    time.sleep(0.5)
                    st.rerun()
        else:
            st.caption("Brak pozycji do uzupełnienia – wszystkie mają opis.")


def main():
    st.title("🚀 CAD Estimator Pro")

    if not init_db():
        st.stop()

    # Sidebar: nawigacja
    st.sidebar.title("Menu")
    
    # ZMIEŃ TĘ LINIĘ:
    # page = st.sidebar.radio("Nawigacja", ["Dashboard", "Nowy projekt", "Historia i Uczenie"])
    
    # NA:
    page = st.sidebar.radio(
        "Nawigacja", 
        ["Dashboard", "Nowy projekt", "Historia i Uczenie", "🛠️ Admin"]
    )
    
    # ... reszta sidebar (bez zmian) ...
    
    # Routing stron
    if page == "Dashboard":
        render_dashboard_page()
    elif page == "Nowy projekt":
        render_new_project_page()
    elif page == "Historia i Uczenie":
        render_history_page()
    elif page == "🛠️ Admin":
        render_admin_page()  # <-- NOWA STRONA
# === MAIN ===
def render_admin_page():
    """Admin panel - zarządzanie danymi"""
    st.header("🛠️ Panel Administratora")
    
    # Auth (prosty protection)
    if "admin_authenticated" not in st.session_state:
        st.session_state["admin_authenticated"] = False
    
    if not st.session_state["admin_authenticated"]:
        password = st.text_input("Hasło administratora", type="password")
        if st.button("Zaloguj"):
            if password == "polmic":  # ZMIEŃ TO NA SWOJE HASŁO!
                st.session_state["admin_authenticated"] = True
                st.rerun()
            else:
                st.error("❌ Błędne hasło")
        st.stop()
    
    # ════════════════════════════════════════════════════════════
    # TABS
    # ════════════════════════════════════════════════════════════
    tab1, tab2, tab3, tab4 = st.tabs([
        "🗂️ Projekty", 
        "🧩 Wzorce (Patterns)", 
        "🔗 Bundles",
        "⚠️ Danger Zone"
    ])
    
    # ════════════════════════════════════════════════════════════
    # TAB 1: PROJEKTY
    # ════════════════════════════════════════════════════════════
    with tab1:
        st.subheader("📋 Zarządzanie projektami")
        
        # Filtry
        col1, col2 = st.columns(2)
        with col1:
            filter_dept = st.selectbox(
                "Dział",
                options=['Wszystkie'] + list(DEPARTMENTS.keys()),
                format_func=lambda x: 'Wszystkie' if x == 'Wszystkie' else f"{x} - {DEPARTMENTS[x]}"
            )
        with col2:
            filter_historical = st.selectbox(
                "Typ",
                ["Wszystkie", "Tylko historyczne", "Tylko bieżące"]
            )
        
        # Pobierz projekty
        with get_db_connection() as conn, conn.cursor(cursor_factory=RealDictCursor) as cur:
            query = "SELECT id, name, client, department, created_at, estimated_hours, actual_hours, is_historical FROM projects WHERE 1=1"
            params = []
            
            if filter_dept != 'Wszystkie':
                query += " AND department = %s"
                params.append(filter_dept)
            
            if filter_historical == "Tylko historyczne":
                query += " AND is_historical = TRUE"
            elif filter_historical == "Tylko bieżące":
                query += " AND (is_historical = FALSE OR is_historical IS NULL)"
            
            query += " ORDER BY created_at DESC LIMIT 100"
            
            cur.execute(query, params)
            projects = cur.fetchall()
        
        if projects:
            st.info(f"Znaleziono {len(projects)} projektów")
            
            # Display in table with delete buttons
            for proj in projects:
                with st.expander(
                    f"[{proj['department']}] {proj['name']} - ID: {proj['id']}" + 
                    (" 📜 HISTORICAL" if proj.get('is_historical') else "")
                ):
                    col_info, col_actions = st.columns([3, 1])
                    
                    with col_info:
                        st.write(f"**Klient:** {proj['client'] or 'N/A'}")
                        st.write(f"**Created:** {proj['created_at'].strftime('%Y-%m-%d %H:%M')}")
                        st.write(f"**Estimated:** {proj['estimated_hours'] or 0:.1f}h")
                        if proj['actual_hours']:
                            st.write(f"**Actual:** {proj['actual_hours']:.1f}h")
                    
                    with col_actions:
                        if st.button("🗑️ Usuń", key=f"del_proj_{proj['id']}", type="secondary"):
                            st.session_state[f"confirm_delete_proj_{proj['id']}"] = True
                        
                        # Confirmation
                        if st.session_state.get(f"confirm_delete_proj_{proj['id']}"):
                            st.warning("⚠️ Na pewno?")
                            col_yes, col_no = st.columns(2)
                            
                            with col_yes:
                                if st.button("✅ TAK", key=f"yes_proj_{proj['id']}"):
                                    with get_db_connection() as conn, conn.cursor() as cur:
                                        # Delete project (cascades to versions)
                                        cur.execute("DELETE FROM projects WHERE id = %s", (proj['id'],))
                                        conn.commit()
                                    st.success(f"✅ Usunięto projekt ID: {proj['id']}")
                                    time.sleep(1)
                                    st.rerun()
                            
                            with col_no:
                                if st.button("❌ NIE", key=f"no_proj_{proj['id']}"):
                                    st.session_state[f"confirm_delete_proj_{proj['id']}"] = False
                                    st.rerun()
        else:
            st.info("Brak projektów spełniających kryteria")
    
    # ════════════════════════════════════════════════════════════
    # TAB 2: WZORCE (PATTERNS)
    # ════════════════════════════════════════════════════════════
    with tab2:
        st.subheader("🧩 Zarządzanie wzorcami komponentów")
        
        # Filtr
        pattern_dept = st.selectbox(
            "Dział",
            options=['Wszystkie'] + list(DEPARTMENTS.keys()),
            format_func=lambda x: 'Wszystkie' if x == 'Wszystkie' else f"{x} - {DEPARTMENTS[x]}",
            key="pattern_dept_filter"
        )
        
        # Pobierz patterns
        with get_db_connection() as conn, conn.cursor(cursor_factory=RealDictCursor) as cur:
            if pattern_dept == 'Wszystkie':
                cur.execute("""
                    SELECT id, name, department, pattern_key, avg_hours_total, 
                           occurrences, confidence, source, last_updated
                    FROM component_patterns
                    ORDER BY department, occurrences DESC
                    LIMIT 200
                """)
            else:
                cur.execute("""
                    SELECT id, name, department, pattern_key, avg_hours_total, 
                           occurrences, confidence, source, last_updated
                    FROM component_patterns
                    WHERE department = %s
                    ORDER BY occurrences DESC
                    LIMIT 200
                """, (pattern_dept,))
            patterns = cur.fetchall()
        
        if patterns:
            st.info(f"Znaleziono {len(patterns)} wzorców")
            
            # Bulk actions
            st.markdown("### 🔧 Akcje grupowe")
            col1, col2, col3 = st.columns(3)
            
            with col1:
                if st.button("🗑️ Usuń wzorce z occ=1", type="secondary"):
                    with get_db_connection() as conn, conn.cursor() as cur:
                        cur.execute("DELETE FROM component_patterns WHERE occurrences <= 1")
                        deleted = cur.rowcount
                        conn.commit()
                    st.success(f"✅ Usunięto {deleted} wzorców")
                    time.sleep(1)
                    st.rerun()
            
            with col2:
                if st.button("🗑️ Usuń bez confidence", type="secondary"):
                    with get_db_connection() as conn, conn.cursor() as cur:
                        cur.execute("DELETE FROM component_patterns WHERE confidence < 0.1")
                        deleted = cur.rowcount
                        conn.commit()
                    st.success(f"✅ Usunięto {deleted} wzorców")
                    time.sleep(1)
                    st.rerun()
            
            with col3:
                min_occ = st.number_input("Min occurrences do zachowania", min_value=1, value=2)
                if st.button(f"🗑️ Usuń < {min_occ} occ", type="secondary"):
                    with get_db_connection() as conn, conn.cursor() as cur:
                        cur.execute("DELETE FROM component_patterns WHERE occurrences < %s", (min_occ,))
                        deleted = cur.rowcount
                        conn.commit()
                    st.success(f"✅ Usunięto {deleted} wzorców")
                    time.sleep(1)
                    st.rerun()
            
            st.markdown("---")
            st.markdown("### 📋 Lista wzorców")
            
            # Individual patterns
            for pat in patterns[:50]:  # Limit display to 50
                with st.expander(
                    f"[{pat['department']}] {pat['name']} - occ: {pat['occurrences']}, conf: {pat['confidence']:.2f}"
                ):
                    col_info, col_del = st.columns([4, 1])
                    
                    with col_info:
                        st.write(f"**ID:** {pat['id']}")
                        st.write(f"**Pattern Key:** {pat['pattern_key']}")
                        st.write(f"**Avg Hours:** {pat['avg_hours_total']:.2f}h")
                        st.write(f"**Source:** {pat['source'] or 'N/A'}")
                        st.write(f"**Updated:** {pat['last_updated'].strftime('%Y-%m-%d %H:%M')}")
                    
                    with col_del:
                        if st.button("🗑️", key=f"del_pat_{pat['id']}"):
                            with get_db_connection() as conn, conn.cursor() as cur:
                                cur.execute("DELETE FROM component_patterns WHERE id = %s", (pat['id'],))
                                conn.commit()
                            st.success("✅ Usunięto")
                            time.sleep(0.5)
                            st.rerun()
            
            if len(patterns) > 50:
                st.info(f"Pokazano 50 z {len(patterns)} wzorców. Użyj filtrów aby zawęzić.")
        
        else:
            st.info("Brak wzorców")
    
    # ════════════════════════════════════════════════════════════
    # TAB 3: BUNDLES
    # ════════════════════════════════════════════════════════════
    with tab3:
        st.subheader("🔗 Zarządzanie bundles (relacje parent→sub)")
        
        bundle_dept = st.selectbox(
            "Dział",
            options=['Wszystkie'] + list(DEPARTMENTS.keys()),
            format_func=lambda x: 'Wszystkie' if x == 'Wszystkie' else f"{x} - {DEPARTMENTS[x]}",
            key="bundle_dept_filter"
        )
        
        with get_db_connection() as conn, conn.cursor(cursor_factory=RealDictCursor) as cur:
            if bundle_dept == 'Wszystkie':
                cur.execute("""
                    SELECT id, department, parent_name, sub_name, 
                           occurrences, total_qty, confidence
                    FROM component_bundles
                    ORDER BY department, occurrences DESC
                    LIMIT 200
                """)
            else:
                cur.execute("""
                    SELECT id, department, parent_name, sub_name, 
                           occurrences, total_qty, confidence
                    FROM component_bundles
                    WHERE department = %s
                    ORDER BY occurrences DESC
                    LIMIT 200
                """, (bundle_dept,))
            bundles = cur.fetchall()
        
        if bundles:
            st.info(f"Znaleziono {len(bundles)} bundles")
            
            # Bulk delete
            min_bundle_occ = st.number_input("Usuń bundles z occ <", min_value=1, value=2)
            if st.button(f"🗑️ Usuń bundles < {min_bundle_occ} occ"):
                with get_db_connection() as conn, conn.cursor() as cur:
                    cur.execute("DELETE FROM component_bundles WHERE occurrences < %s", (min_bundle_occ,))
                    deleted = cur.rowcount
                    conn.commit()
                st.success(f"✅ Usunięto {deleted} bundles")
                time.sleep(1)
                st.rerun()
            
            st.markdown("---")
            
            for bundle in bundles[:50]:
                with st.expander(
                    f"[{bundle['department']}] {bundle['parent_name']} → {bundle['sub_name']} (occ: {bundle['occurrences']})"
                ):
                    col_i, col_d = st.columns([4, 1])
                    with col_i:
                        st.write(f"**Avg Qty:** {bundle['total_qty'] / bundle['occurrences']:.1f}")
                        st.write(f"**Confidence:** {bundle['confidence']:.2f}")
                    with col_d:
                        if st.button("🗑️", key=f"del_bun_{bundle['id']}"):
                            with get_db_connection() as conn, conn.cursor() as cur:
                                cur.execute("DELETE FROM component_bundles WHERE id = %s", (bundle['id'],))
                                conn.commit()
                            st.success("✅")
                            time.sleep(0.5)
                            st.rerun()
        else:
            st.info("Brak bundles")
    
    # ════════════════════════════════════════════════════════════
    # TAB 4: DANGER ZONE
    # ════════════════════════════════════════════════════════════
    with tab4:
        st.subheader("⚠️ DANGER ZONE - Operacje nieodwracalne")
        
        st.error("⚠️ Te operacje są NIEODWRACALNE! Nie ma backup automatycznego!")
        
        with st.expander("🗑️ Usuń WSZYSTKIE projekty z działu"):
            danger_dept = st.selectbox(
                "Wybierz dział do wyczyszczenia",
                options=list(DEPARTMENTS.keys()),
                format_func=lambda x: f"{x} - {DEPARTMENTS[x]}",
                key="danger_dept"
            )
            
            confirmation = st.text_input(
                f"Wpisz '{danger_dept}' aby potwierdzić",
                key="danger_confirm_dept"
            )
            
            if st.button("🗑️ USUŃ WSZYSTKIE PROJEKTY", type="secondary"):
                if confirmation == danger_dept:
                    with get_db_connection() as conn, conn.cursor() as cur:
                        cur.execute("DELETE FROM projects WHERE department = %s", (danger_dept,))
                        deleted = cur.rowcount
                        conn.commit()
                    st.success(f"✅ Usunięto {deleted} projektów z działu {danger_dept}")
                else:
                    st.error("❌ Błędne potwierdzenie!")
        
        with st.expander("🗑️ Usuń WSZYSTKIE wzorce z działu"):
            danger_dept2 = st.selectbox(
                "Wybierz dział",
                options=list(DEPARTMENTS.keys()),
                format_func=lambda x: f"{x} - {DEPARTMENTS[x]}",
                key="danger_dept2"
            )
            
            confirmation2 = st.text_input(
                f"Wpisz '{danger_dept2}' aby potwierdzić",
                key="danger_confirm_dept2"
            )
            
            if st.button("🗑️ USUŃ WSZYSTKIE WZORCE", type="secondary"):
                if confirmation2 == danger_dept2:
                    with get_db_connection() as conn, conn.cursor() as cur:
                        cur.execute("DELETE FROM component_patterns WHERE department = %s", (danger_dept2,))
                        deleted = cur.rowcount
                        conn.commit()
                    st.success(f"✅ Usunięto {deleted} wzorców z działu {danger_dept2}")
                else:
                    st.error("❌ Błędne potwierdzenie!")
        
        with st.expander("🗑️ RESET CAŁEJ BAZY (wszystko)"):
            st.error("⚠️⚠️⚠️ TO USUNIE WSZYSTKO! ⚠️⚠️⚠️")
            
            confirm_reset = st.text_input("Wpisz 'DELETE EVERYTHING' aby potwierdzić")
            
            if st.button("💣 RESET DATABASE", type="secondary"):
                if confirm_reset == "DELETE EVERYTHING":
                    with get_db_connection() as conn, conn.cursor() as cur:
                        cur.execute("TRUNCATE TABLE component_bundles CASCADE")
                        cur.execute("TRUNCATE TABLE category_baselines CASCADE")
                        cur.execute("TRUNCATE TABLE project_versions CASCADE")
                        cur.execute("TRUNCATE TABLE component_patterns CASCADE")
                        cur.execute("TRUNCATE TABLE projects CASCADE")
                        conn.commit()
                    st.success("✅ Baza wyczyszczona całkowicie!")
                    time.sleep(2)
                    st.rerun()
                else:
                    st.error("❌ Błędne potwierdzenie!")


# ════════════════════════════════════════════════════════════
# ZAKTUALIZUJ MAIN() - dodaj zakładkę Admin
# ════════════════════════════════════════════════════════════
def main():
    st.title("🚀 CAD Estimator Pro")

    if not init_db():
        st.stop()

    # Sidebar: nawigacja
    st.sidebar.title("Menu")
    
    page = st.sidebar.radio("Nawigacja", ["Dashboard", "Nowy projekt", "Historia i Uczenie", "🛠️ Admin"])

    # Sidebar: ustawienia AI
    st.sidebar.subheader("Ustawienia AI")
    
    # 1) Model tekstowy (dla estymacji, JSON)
    available_text_models = [
        m for m in list_local_models() 
        if not any(m.startswith(p) for p in ("llava", "bakllava", "moondream", "qwen2-vl", "qwen2.5vl", "nomic-embed"))
    ]
    
    if "selected_text_model" not in st.session_state:
        # Preferuj qwen2.5 dla technicznego tekstu
        default_text = "qwen2.5:7b" if "qwen2.5:7b" in available_text_models else (
            "mistral:7b-instruct" if "mistral:7b-instruct" in available_text_models else 
            (available_text_models[0] if available_text_models else "llama3:latest")
        )
        st.session_state["selected_text_model"] = default_text
    
    try:
        text_idx = available_text_models.index(st.session_state["selected_text_model"])
    except (ValueError, IndexError):
        text_idx = 0
    
    selected_text_model = st.sidebar.selectbox(
        "Model AI (estymacja/JSON)",
        options=available_text_models or ["llama3:latest"],
        index=text_idx,
        key="text_model_sel",
        help="Model do analizy komponentów, generowania JSON, opisów zadań"
    )
    st.session_state["selected_text_model"] = selected_text_model
    
    # 2) Model Vision (dla obrazów/rysunków)
    available_vision_models = [
        m for m in list_local_models()
        if any(m.startswith(p) for p in ("llava", "bakllava", "moondream", "qwen2-vl", "qwen2.5vl"))
    ]
    
    if available_vision_models:
        if "selected_vision_model" not in st.session_state:
            # Preferuj qwen2.5vl dla technicznych rysunków
            default_vision = "qwen2.5vl:7b" if "qwen2.5vl:7b" in available_vision_models else (
                "qwen2-vl:7b" if "qwen2-vl:7b" in available_vision_models else 
                available_vision_models[0]
            )
            st.session_state["selected_vision_model"] = default_vision
        
        try:
            vision_idx = available_vision_models.index(st.session_state["selected_vision_model"])
        except (ValueError, IndexError):
            vision_idx = 0
        
        selected_vision_model = st.sidebar.selectbox(
            "Model Vision (obrazy/rysunki)",
            options=available_vision_models,
            index=vision_idx,
            key="vision_model_sel",
            help="Model do analizy zdjęć, schematów, rysunków technicznych"
        )
        st.session_state["selected_vision_model"] = selected_vision_model
    else:
        st.sidebar.warning("⚠️ Brak modeli Vision (zainstaluj llava/qwen2-vl)")
        selected_vision_model = None

    st.sidebar.markdown("---")
    st.sidebar.subheader("🌐 Web Lookup")
    
    if "allow_web_lookup" not in st.session_state:
        st.session_state["allow_web_lookup"] = False
    
    allow_web = st.sidebar.checkbox(
        "Zezwól na web lookup (normy/benchmarki)",
        value=st.session_state["allow_web_lookup"],
        key="web_lookup_toggle",
        help="Pobiera publiczne dane: normy ISO/EN, benchmarki czasów, dostępność komponentów. NIE wysyła danych projektu!"
    )
    st.session_state["allow_web_lookup"] = allow_web
    
    if allow_web:
        st.sidebar.caption("✅ Web lookup aktywny - system może wzbogacić estymację o dane z sieci")
    else:
        st.sidebar.caption("🔒 Tryb offline - tylko lokalne wzorce")

    st.sidebar.subheader("Status Systemu")
    st.sidebar.write(f"Ollama AI: {'✅ Połączony' if any(list_local_models()) else '❌ Brak połączenia'}")

    with st.sidebar.expander("Dostępne modele"):
        models = list_local_models()
        if models:
            st.write("\n".join(f"- `{m}`" for m in models))
        else:
            st.write("Brak modeli")

    st.sidebar.markdown("---")
    if st.sidebar.button("🔄 Odśwież listę modeli"):
        try:
            list_local_models.cache_clear()
        except Exception:
            pass
        st.rerun()

    st.sidebar.subheader("Embedding (diagnostyka)")
    detected_dim = detect_embed_dim(EMBED_MODEL)
    if detected_dim and detected_dim != EMBED_DIM:
        st.sidebar.error(f"EMBED_DIM={EMBED_DIM} vs model '{EMBED_MODEL}' zwraca {detected_dim}. Zmień EMBED_DIM lub model.")
    elif detected_dim:
        st.sidebar.success(f"Model '{EMBED_MODEL}' OK (dim={detected_dim}).")
    else:
        st.sidebar.info("Nie udało się pobrać embeddingu (sprawdź OLLAMA_URL / model).")

    # Demo/próbne dane
    with st.sidebar.expander("🧪 Demo / Próbne dane", expanded=False):
        if st.button("Wypełnij formularz przykładowymi danymi"):
            fill_demo_fields()
        demo_excel = generate_sample_excel()
        st.download_button("📥 Pobierz przykładowy Excel", demo_excel,
                           file_name="demo_estymacja.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        demo_pdf = generate_sample_pdf()
        if demo_pdf:
            st.download_button("📥 Pobierz przykładowy PDF", demo_pdf, file_name="demo_spec.pdf", mime="application/pdf")
        else:
            st.info("Aby generować PDF, zainstaluj: pip install reportlab")
        demo_img = generate_sample_image()
        st.download_button("📥 Pobierz przykładowy obraz (PNG)", demo_img, file_name="demo_schemat.png", mime="image/png")

    # Routing stron
    if page == "Dashboard":
        render_dashboard_page()

    elif page == "Nowy projekt":
        render_new_project_page()
    
    elif page == "Historia i Uczenie":
        render_history_page()

    elif page == "🛠️ Admin":
        render_admin_page()  # <-- NOWA STRONA

if __name__ == "__main__":

# CADEstimator_final.py
# DODAJ NA KOŃCU PLIKU (przed main())


    main()

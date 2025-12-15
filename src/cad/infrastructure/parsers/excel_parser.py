"""
CAD Estimator Pro - Excel Parser

Implementation of ExcelParser protocol.
"""
import logging
from io import BytesIO
from typing import BinaryIO
import pandas as pd
from openpyxl import load_workbook

from cad.domain.exceptions import ExcelParsingError

logger = logging.getLogger(__name__)


class CADExcelParser:
    """
    Excel parser for CAD project files.

    Parses component hierarchy, hours, and comments from Excel files.
    """

    def __init__(self):
        """Initialize ExcelParser."""
        self.COL_POS = 0
        self.COL_DESC = 1
        self.COL_COMMENT = 2
        self.COL_STD_PARTS = 3
        self.COL_SPEC_PARTS = 4
        self.COL_HOURS_LAYOUT = 7
        self.COL_HOURS_DETAIL = 9
        self.COL_HOURS_DOC = 11
        self.DATA_START_ROW = 11

    def parse(self, file: BinaryIO) -> dict:
        """Parse Excel file."""
        try:
            content = file.read() if hasattr(file, 'read') else file
            df = pd.read_excel(BytesIO(content), header=None)

            # Parse multipliers
            multipliers = self._parse_multipliers(df)

            # Parse components
            components = []
            for row_idx in range(self.DATA_START_ROW, df.shape[0]):
                try:
                    comp = self._parse_component_row(df, row_idx)
                    if comp:
                        components.append(comp)
                except Exception as e:
                    logger.warning(f"Failed to parse row {row_idx + 1}: {e}")
                    continue

            # Calculate totals
            parts_only = [c for c in components if not c.get('is_summary', False)]
            totals = {
                'layout': sum(c.get('hours_3d_layout', 0) for c in parts_only),
                'detail': sum(c.get('hours_3d_detail', 0) for c in parts_only),
                'documentation': sum(c.get('hours_2d', 0) for c in parts_only)
            }
            totals['total'] = sum(totals.values())

            statistics = {
                'parts_count': len(parts_only),
                'assemblies_count': sum(1 for c in components if c.get('is_summary', False))
            }

            logger.info(f"✅ Parsed Excel: {len(parts_only)} components, {totals['total']:.1f}h total")

            return {
                'components': components,
                'totals': totals,
                'multipliers': multipliers,
                'statistics': statistics
            }

        except Exception as e:
            logger.error(f"Excel parsing failed: {e}", exc_info=True)
            raise ExcelParsingError(f"Failed to parse Excel file: {e}")

    def extract_description_from_a1(self, file: BinaryIO) -> str:
        """Extract project description from cell A1 (first sheet)."""
        try:
            content = file.read() if hasattr(file, 'read') else file
            wb = load_workbook(BytesIO(content), data_only=True)
            ws = wb.worksheets[0]
            val = ws.cell(row=1, column=1).value
            return str(val).strip() if val is not None else ""
        except Exception as e:
            logger.info(f"Could not read A1 from Excel: {e}")
            return ""

    def _parse_multipliers(self, df: pd.DataFrame) -> dict:
        """Parse multipliers from row 10 (index 9)."""
        try:
            return {
                'layout': float(df.iloc[9, self.COL_HOURS_LAYOUT]) if pd.notna(df.iloc[9, self.COL_HOURS_LAYOUT]) else 1.0,
                'detail': float(df.iloc[9, self.COL_HOURS_DETAIL]) if pd.notna(df.iloc[9, self.COL_HOURS_DETAIL]) else 1.0,
                'documentation': float(df.iloc[9, self.COL_HOURS_DOC]) if pd.notna(df.iloc[9, self.COL_HOURS_DOC]) else 1.0
            }
        except Exception:
            return {'layout': 1.0, 'detail': 1.0, 'documentation': 1.0}

    def _parse_component_row(self, df: pd.DataFrame, row_idx: int) -> dict | None:
        """Parse single component row."""
        pos = str(df.iloc[row_idx, self.COL_POS]).strip() if pd.notna(df.iloc[row_idx, self.COL_POS]) else ""
        name = str(df.iloc[row_idx, self.COL_DESC]).strip() if pd.notna(df.iloc[row_idx, self.COL_DESC]) else ""

        if not pos or pos in ['nan', 'None', '']:
            return None

        if not name or name in ['nan', 'None']:
            name = f"[Pozycja {pos}]"

        comment = str(df.iloc[row_idx, self.COL_COMMENT]).strip() if pd.notna(df.iloc[row_idx, self.COL_COMMENT]) else ""

        hours_layout = float(df.iloc[row_idx, self.COL_HOURS_LAYOUT]) if pd.notna(df.iloc[row_idx, self.COL_HOURS_LAYOUT]) else 0.0
        hours_detail = float(df.iloc[row_idx, self.COL_HOURS_DETAIL]) if pd.notna(df.iloc[row_idx, self.COL_HOURS_DETAIL]) else 0.0
        hours_doc = float(df.iloc[row_idx, self.COL_HOURS_DOC]) if pd.notna(df.iloc[row_idx, self.COL_HOURS_DOC]) else 0.0

        is_summary = bool(pos.count(',') == 1 and pos.endswith(',0')) or pos.isdigit()

        return {
            'id': pos,
            'name': name,
            'comment': comment,
            'hours_3d_layout': hours_layout,
            'hours_3d_detail': hours_detail,
            'hours_2d': hours_doc,
            'hours': hours_layout + hours_detail + hours_doc,
            'is_summary': is_summary
        }
